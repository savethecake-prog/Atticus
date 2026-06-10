"""One runnable suite for the whole skill. Run it before trusting any new check,
decoder, or range:

    cd salt-input-builder && python tests/run_tests.py

It exits non-zero if anything fails. It consolidates the per-module guards and
runs the engine fixture cases in tests/fixtures/. Network-dependent link
calibration is covered offline here (the soft-404 detector); the live calibration
is exercised separately when a job runs.
"""
import os, sys, glob

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, os.path.join(ROOT, "scripts"))

import yaml
from openpyxl import Workbook
import profile_engine, profile_builder, ledger, linkcheck

PASS, FAIL = [], []


def check(name, cond, detail=""):
    (PASS if cond else FAIL).append(name + (f"  [{detail}]" if detail and not cond else ""))


def _sheet(roles, labels, row):
    wb = Workbook(); ws = wb.active
    for c, v in row.items():
        ws.cell(1, int(c), v)
    return ws


# --- engine fixture cases ---------------------------------------------------
def test_engine_fixtures():
    for path in sorted(glob.glob(os.path.join(HERE, "fixtures", "*.yaml"))):
        for case in yaml.safe_load(open(path)):
            roles = {k: int(v) if str(v).isdigit() else v for k, v in case["roles"].items()}
            ws = _sheet(roles, case.get("labels", {}), {int(k): v for k, v in case["row"].items()})
            found = profile_engine.run_checks(ws, 1, roles, case.get("labels", {}), case["profile"])
            exp = case.get("expect", [])
            ok = len(found) == len(exp) and all(
                any(e["sev"] == sev and e["contains"] in msg for _, sev, msg in found) for e in exp
            )
            check(f"engine fixture: {case['name']}", ok, f"got {[(s, m) for _, s, m in found]}")


# --- engine number-parsing guard --------------------------------------------
def test_engine_guards():
    check("engine: tolerance (+-10%) ignored", profile_engine._nums("Up to 2000 RPM (+-10%)") == [2000.0])
    check("engine: range min/max kept", profile_engine._nums("400 (+-200) - 1200 RPM (+-10%)") == [400.0, 1200.0])


# --- conflict comparator branches -------------------------------------------
def test_conflicts():
    man = lambda v: {"value": v, "provenance": "manufacturer", "source_url": "https://m", "snippet": v}
    ret = lambda v, u="https://s": {"value": v, "provenance": "retailer", "source_url": u, "snippet": v}
    E = lambda val, cands, prov="manufacturer": {"tab": "T", "row": 2, "field": "f", "column": 5,
                                                  "value": val, "provenance": prov, "source_url": "https://m",
                                                  "snippet": val, "candidates": cands}
    sev = lambda e: (ledger.detect_conflicts([e]) or [(0, 0, 0, 0, None)])[0][4]
    check("conflict: same-tier numbers disagree -> HARD", sev(E("135 g", [ret("135 g", "https://a"), ret("120 g", "https://b")], "retailer")) == "HARD")
    check("conflict: cross-tier, chose manufacturer -> SOFT", sev(E("120 g", [man("120 g"), ret("135 g")])) == "SOFT")
    check("conflict: chose lower tier -> HARD", sev(E("135 g", [man("120 g"), ret("135 g")], "retailer")) == "HARD")
    check("conflict: wording-only differs -> SOFT", sev(E("HDB", [man("HDB"), ret("hydro dynamic")])) == "SOFT")
    check("conflict: agree after norm -> none", sev(E("120 g", [man("120 g"), ret("120g")])) is None)


# --- ledger validation of candidates ----------------------------------------
def test_ledger_validate():
    man = lambda v, s: {"value": v, "provenance": "manufacturer", "source_url": "https://m", "snippet": s}
    bad_choice = {"tab": "T", "row": 2, "field": "f", "column": 5, "value": "999",
                  "provenance": "manufacturer", "source_url": "https://m", "snippet": "999",
                  "candidates": [man("120", "120"), man("135", "135")]}
    check("validate: chosen not among candidates caught",
          any("not among its candidates" in v for v in ledger.validate([bad_choice])))
    bad_snip = {"tab": "T", "row": 2, "field": "f", "column": 5, "value": "120",
                "provenance": "manufacturer", "source_url": "https://m", "snippet": "heavy",
                "candidates": [man("120", "weight is heavy")]}
    check("validate: non-verbatim value without a doubt_reason is caught",
          any("not verbatim" in v for v in ledger.validate([bad_snip])))
    # the key new behaviour: a doubt_reason lets an existing-but-unverbatim value through
    disputed = {"tab": "T", "row": 3, "field": "tc", "column": 6, "value": "12.5 W/mK",
                "provenance": "manufacturer", "source_url": "https://m",
                "snippet": "thermal paste for high-end CPUs",
                "doubt_reason": "long-published figure; not on the current manufacturer datasheet - confirm or remove"}
    check("validate: doubt_reason lets a non-verbatim value through (no omission)",
          ledger.validate([disputed]) == [])
    import confidence as C
    s, r = C.score_entry(disputed)
    check("confidence: disputed value scored low and carries its doubt reason",
          s == 30 and "current manufacturer datasheet" in r and not C.eligible(disputed))


# --- profile builder inference ----------------------------------------------
def test_profile_builder():
    products = [
        {"family": "SC140", "sku": "3MSC14FA1.0001", "attrs": {"blade": "forward", "colour": "black"}},
        {"family": "SC140", "sku": "3MSC14RA1.0001", "attrs": {"blade": "reverse", "colour": "black"}},
        {"family": "SC140", "sku": "3MSC14FA1.0002", "attrs": {"blade": "forward", "colour": "white"}},
        {"family": "SC140", "sku": "3MSC14RA1.0002", "attrs": {"blade": "reverse", "colour": "white"}},
        {"family": "Unity", "sku": "3MUN24FA1.0001", "attrs": {"size": "240", "blade": "forward", "colour": "black"}},
        {"family": "Unity", "sku": "3MUN36RA1.0002", "attrs": {"size": "360", "blade": "reverse", "colour": "white"}},
    ]
    prof, _ = profile_builder.build_profile("fans", "Cougar", {"5": "Product colour"}, products)
    names = {d["name"] for d in prof["decoders"]}
    sc = next((d for d in prof["decoders"] if d["name"] == "SC140"), None)
    check("builder: SC140 decoder inferred", sc is not None)
    check("builder: SC140 blade map correct", bool(sc) and sc.get("blade_map") == {"F": "forward", "R": "reverse"})
    check("builder: SC140 suffix colour correct", bool(sc) and sc.get("suffix_colour") == {"1": "black", "2": "white"})
    check("builder: confounded Unity NOT guessed", "Unity" not in names)
    import re as _re
    check("builder: inferred regex matches its own SKUs",
          bool(sc) and all(_re.search(sc["sku_regex"], p["sku"]) for p in products if p["family"] == "SC140"))


# --- linkcheck soft-404 (offline) -------------------------------------------
def test_linkcheck_offline():
    notfound = "<html><head><title>404 Not Found</title></head><body>gone</body></html>"
    normal = "<html><head><title>Apolar 120 ARGB Fan</title></head><body>specs...</body></html>"
    check("linkcheck: soft-404 title detected", linkcheck._looks_like_404(notfound) is True)
    check("linkcheck: normal page not a 404", linkcheck._looks_like_404(normal) is False)


# --- duplicate identifier across rows (agnostic core audit) -----------------
def test_duplicate_sku():
    import audit
    wb = Workbook(); ws = wb.active; ws.title = "Fans"
    for c, h in {1: "Title", 2: "Spec", 3: "SKU"}.items():
        ws.cell(1, c, h)
    ws.cell(2, 1, "A"); ws.cell(2, 3, "DUP-1")
    ws.cell(3, 1, "B"); ws.cell(3, 3, "DUP-1")
    os.makedirs("/tmp/th", exist_ok=True); wb.save("/tmp/th/wb.xlsx")
    schema = {"tabs": {"Fans": {"roles": {"title": 1, "spec_start": 2, "spec_end": 2, "sku": 3},
                                "labels": {}, "header_row": 1, "first_data_row": 2, "example_rows": []}}}
    ledger.save("/tmp/th/led.json", [])
    findings = audit.audit("/tmp/th/wb.xlsx", schema, "/tmp/th/led.json", "/tmp/th/rep.md", profiles={})
    check("audit: duplicate SKU across products flagged",
          any("duplicate SKU" in m for *_, m in findings))


def test_jobspec():
    import jobspec
    good = {"job": {"name": "B1", "template": "t.xlsx", "output_dir": "out"},
            "source_policy": {"tier_order": ["manufacturer", "retailer"]},
            "authority": {"infer_brand_domain": True},
            "confidence_colours": {"manufacturer": {"fill": "C6EFCE", "font": "006100"}},
            "category_profiles": {"Fans": "profiles/fans.yaml"},
            "gates": {"mode": "adaptive"}, "deliverables": ["workbook"],
            "items": [{"tab": "Fans", "client_fields": {"sku": "X-1"}}]}
    check("jobspec: valid spec passes", jobspec.is_valid(good))
    bad = {"job": {}, "source_policy": {"tier_order": ["wholesaler"]},
           "gates": {"mode": "nope"}, "category_profiles": [], "items": "nope"}
    check("jobspec: malformed spec rejected", not jobspec.is_valid(bad))


def test_schema_store():
    import schema_store, tempfile
    d = tempfile.mkdtemp()
    schema_store.save(d, {"tabs": {"Fans": {"signature": "sig1", "roles": {"title": 1}, "labels": {}}}})
    check("schema_store: find by signature", bool(schema_store.find_by_signature(d, "sig1")))
    s2 = {"tabs": {"Fans": {"signature": "sig1"}, "Mobo": {"signature": "new"}}}
    n = schema_store.annotate(d, s2)
    check("schema_store: reuse on match, confirm on new",
          n == 1 and s2["tabs"]["Mobo"]["needs_confirmation"] and s2["tabs"]["Fans"]["reused"])


def test_runner():
    import runner
    schema = {"tabs": {"Fans": {"roles": {"title": 1, "spec_start": 2, "spec_end": 2, "sku": 3},
                                "labels": {}, "header_row": 1, "first_data_row": 2, "example_rows": []}}}
    ret = lambda v, u: {"value": v, "provenance": "retailer", "source_url": u, "snippet": v}
    entries = [{"tab": "Fans", "row": 2, "field": "noise", "column": 2, "value": "26 dBA",
                "provenance": "retailer", "source_url": "https://a", "snippet": "26 dBA",
                "candidates": [ret("26 dBA", "https://a"), ret("31 dBA", "https://b")]}]
    q = runner.needs_you(schema, entries)
    check("runner: needs_you surfaces hard conflict", any("conflict" in n for n in q))
    check("runner: delivery summary renders", "Delivery summary" in runner.delivery_summary(schema, entries))
    check("runner: status on empty dir -> contract phase", "contract" in runner.status("/tmp/no-such-job-xyz")["next"])


def test_primary_source():
    E = lambda prov, url, ok: {"tab": "T", "row": 2, "field": "f", "column": 3,
                               "value": "v", "provenance": prov, "source_url": url, "link_ok": ok}
    man = "https://maker"; ret = "https://shop"; az = "https://amazon"; ds = "https://mirror"
    # manufacturer live beats retailer live (authority)
    check("picker: manufacturer-live preferred",
          ledger.primary_source([E("manufacturer", man, True), E("retailer", ret, True)], "T", 2) == man)
    # the bug: manufacturer DEAD must lose to a live retailer
    check("picker: dead manufacturer loses to live retailer",
          ledger.primary_source([E("manufacturer", man, False), E("retailer", ret, True)], "T", 2) == ret)
    # unverified manufacturer still beats live retailer (only dead is demoted)
    check("picker: unverified manufacturer keeps authority over live retailer",
          ledger.primary_source([E("manufacturer", man, None), E("retailer", ret, True)], "T", 2) == man)
    # Aeronaut shape: amazon dead, mirror live, titanrig unverified -> live mirror
    check("picker: chooses the live link among retailers (Aeronaut case)",
          ledger.primary_source([E("retailer", az, False), E("retailer", ds, True), E("retailer", ret, None)], "T", 2) == ds)
    # all dead -> still returns the best-tier (flagged elsewhere)
    check("picker: all dead returns best tier",
          ledger.primary_source([E("retailer", ret, False), E("manufacturer", man, False)], "T", 2) == man)


def test_confidence():
    import confidence as C
    ret = lambda v, u: {"value": v, "provenance": "retailer", "source_url": u, "snippet": v}
    E = lambda val, prov, cands=None, url=None: {"value": val, "provenance": prov, "source_url": url, "snippet": val, "candidates": cands}
    check("confidence: manufacturer single = 90 and eligible",
          C.score_entry(E("x", "manufacturer", url="https://m"))[0] == 90 and C.eligible(E("x", "manufacturer", url="https://m")))
    check("confidence: single retailer = 40 and NOT eligible (2-source rule)",
          C.score_entry(E("x", "retailer", url="https://a"))[0] == 40 and not C.eligible(E("x", "retailer", url="https://a")))
    two = E("x", "retailer", cands=[ret("x", "https://a"), ret("x", "https://b")])
    check("confidence: two independent retailers = 70 and eligible", C.score_entry(two)[0] == 70 and C.eligible(two))
    same = E("x", "retailer", cands=[ret("x", "https://a/1"), ret("x", "https://a/2")])
    check("confidence: same host counts once (not eligible)", not C.eligible(same))
    dis = E("x", "retailer", cands=[ret("x", "https://a"), ret("y", "https://b")])
    check("confidence: disagreement = 35", C.score_entry(dis)[0] == 35)
    check("confidence: client identity = 85 and eligible",
          C.score_entry(E("TG-1", "client"))[0] == 85 and C.eligible(E("TG-1", "client")))
    check("confidence: bands map correctly (high/medium/low)",
          C.band(90) == "high" and C.band(70) == "medium" and C.band(40) == "low")
    check("confidence: single-source reason names the host",
          "a.com" in C.score_entry(E("x", "retailer", url="https://a.com"))[1])


def test_fields():
    import fields
    tab = {"roles": {"spec_start": 3, "spec_end": 25, "header_row": 1},
           "labels": {"3": "Electrical conductivity:", "4": "Supported application:", "5": "Components:",
                      "19": "Colour:", "20": "Operating temperature: "}}
    dec = {d["field"]: (d["action"], d["col"]) for d in fields.reconcile(tab,
           [{"field": "Electrical conductivity"}, {"field": "Thermal conductivity (W/mK)"},
            {"field": "Viscosity (Pa.s)"}, {"field": "Density (g/cm3)"}])}
    check("fields: known field maps to its existing column",
          dec["Electrical conductivity"] == ("existing", 3))
    check("fields: 'thermal conductivity' does NOT collide with 'electrical conductivity'",
          dec["Thermal conductivity (W/mK)"][0] == "relabel")
    check("fields: new fields fill blank spare columns in range",
          dec["Viscosity (Pa.s)"][0] == "relabel" and dec["Density (g/cm3)"][0] == "relabel")
    # exact vs non-identical existing match
    d2 = fields.reconcile(tab, [{"field": "Operating temperature"}, {"field": "Operating temperature (C)"}])
    by2 = {x["field"]: x for x in d2}
    check("fields: identical header match flagged exact",
          by2["Operating temperature"]["action"] == "existing" and by2["Operating temperature"]["exact"] is True)
    check("fields: non-identical header match still fills, marked not exact",
          by2["Operating temperature (C)"]["action"] == "existing" and by2["Operating temperature (C)"]["exact"] is False)


def test_hide_blank_columns():
    from openpyxl import Workbook
    import build
    wb = Workbook(); ws = wb.active
    ws.cell(1, 1, "Title"); ws.cell(2, 1, "P1"); ws.cell(3, 1, "P2")
    ws.cell(1, 3, "A"); ws.cell(1, 4, "B")
    ws.cell(2, 3, "value")  # col 3 has data, col 4 entirely blank
    t = {"roles": {"title": 1, "spec_start": 3, "spec_end": 4}, "header_row": 1}
    n = build._hide_blank_columns(ws, t)
    check("hide: one fully-blank column hidden", n == 1)
    check("hide: blank column hidden, populated column visible",
          ws.column_dimensions["D"].hidden and not ws.column_dimensions["C"].hidden)


def test_from_spec_table():
    import fields
    raw = "Power output: 750 W\nEfficiency: 80 Plus Gold\nModular: Fully modular\nFan: 120mm"
    out = fields.from_spec_table(raw)
    names = [d["field"] for d in out]
    check("from_spec_table: one field per row", len(out) == 4, str(names))
    check("from_spec_table: strips value", "Power output" in names)
    d = fields.from_spec_table({"Wattage": "750", "Fan": "120mm", "Wattage ": "dup"})
    check("from_spec_table: dedup by normalised label", len(d) == 2, str([x["field"] for x in d]))
    tab = fields.from_spec_table(["Length\t160mm", "Width\t150mm"])
    check("from_spec_table: tab-separated", [x["field"] for x in tab] == ["Length", "Width"])


def test_completeness():
    import completeness
    wb = Workbook()
    ws = wb.active; ws.title = "Fans"
    heads = ["Title", "Size", "Speed", "Airflow", "Noise", "Bearing", "Brand:", "SKU:"]
    for c, h in enumerate(heads, 1):
        ws.cell(1, c, h)
    # rich row: 5 specs; thin row: 1 spec
    ws.append(["Fan A", "120mm", "2000", "68 CFM", "28 dBA", "FDB", "Cougar", "SKU1"])
    ws.append(["Fan B", "120mm", None, None, None, None, "Cougar", "SKU2"])
    ws.append(["Example fan", "x", "y", "z", "w", "v", "Cougar", "SKU3"])  # ignored
    path = os.path.join(HERE, "_tmp_completeness.xlsx"); wb.save(path)
    schema = {"tabs": {"Fans": {}}}
    f = completeness.check(path, schema, ratio=0.6, floor=4)
    rows = {x["row"]: x for x in f if x["row"]}
    check("completeness: thin row flagged", 3 in rows, str(rows))
    check("completeness: rich row not flagged", 2 not in rows)
    check("completeness: example row ignored", all(x["product"] != "Example fan" for x in f))
    # source_field_count catches systemic thinness (rich row's 5 < ratio*10)
    schema["tabs"]["Fans"]["source_field_count"] = 10
    f2 = completeness.check(path, schema, ratio=0.6, floor=4)
    rows2 = {x["row"] for x in f2 if x["row"]}
    check("completeness: source count flags the rich row too", 2 in rows2, str(rows2))
    tabflag = [x for x in f2 if x["row"] is None]
    check("completeness: under-discovered tab flagged", len(tabflag) == 1, str(tabflag))
    os.remove(path)


# --- identity matching (Phase A: regression from the Infinity 2026-06-09 failures) ---
def test_match():
    import match
    T = [
        {"id": "OA40SBB", "brand": "Oppo", "sku": "OA40SBB", "barcode": "",
         "title": "Oppo A40 (Sparkle Black, 128 GB) (4 GB RAM)"},
        {"id": "OA40SBL", "brand": "Oppo", "sku": "OA40SBL", "barcode": "",
         "title": "Oppo A40 Starry Blue (Starry Purple, 128 GB) (4 GB RAM)"},
        {"id": "OR12FOG", "brand": "Oppo", "sku": "OR12FOG", "barcode": "",
         "title": "oppo Reno12 F 4G Dual Sim 256GB (Green, 256 GB) (8 GB RAM) +Free Oppo Jaua Watch"},
        {"id": "OR12FMG", "brand": "Oppo", "sku": "OR12FMG", "barcode": "",
         "title": "Oppo Reno12 F (Matte Grey, 256 GB) (8 GB RAM)+Free Oppo Jaua Watch"},
        {"id": "CPH2629", "brand": "Oppo", "sku": "CPH2629", "barcode": "",
         "title": "Oppo Reno12 Pro 5G (Space Brown, 512 GB) (12 GB RAM) +Free Oppo Jaua Watch"},
        {"id": "OR13BL", "brand": "Oppo", "sku": "OR13BL", "barcode": "",
         "title": "Oppo Reno 13 5G (Luminous Blue, 512 GB) (12 GB RAM)+Free Oppo Jaua Watch"},
        {"id": "AOPROG", "brand": "Oppo", "sku": "AOPROG", "barcode": "",
         "title": "Oppo A5 Pro (Olive Green, 256 GB) (8 GB RAM)"},
        {"id": "OA5MP", "brand": "Oppo", "sku": "OA5MP", "barcode": "",
         "title": "Oppo A5 (Midnight Purple, 256 GB) (8 GB RAM)"},
        {"id": "H72B", "brand": "Hisense", "sku": "H72B", "barcode": "",
         "title": "Hisense H72 with Free Bluetooth Speaker"},
        {"id": "H72BLL", "brand": "Hisense", "sku": "H72BLL", "barcode": "",
         "title": "Hisense H72 (Blue, 128 GB)"},
        {"id": "V30E", "brand": "Vivo", "sku": "6935117884578", "barcode": "6935117884578",
         "title": "vivo V30e (12 GB RAM)"},
    ]
    S = [
        {"id": "a40sb", "brand": "Oppo", "colour": "Sparkle Black", "storage": "128",
         "model": "Oppo A40 Sparkle Black 4GB RAM 128", "model_number": "CPH2669"},
        {"id": "a40sp", "brand": "Oppo", "colour": "Starry Purple", "storage": "128",
         "model": "Oppo A40 Starry Purple 4GB RAM 128", "model_number": "CPH2669"},
        {"id": "r12fog", "brand": "Oppo", "colour": "Olive Green", "storage": "256",
         "model": "Oppo Reno 12F Olive Green 8GB RAM 256", "model_number": "CPH2687"},
        {"id": "r12fmg", "brand": "Oppo", "colour": "Matte Gray", "storage": "256",
         "model": "Oppo Reno 12 F Matte Gray 8GB RAM 256", "model_number": "CPH2687"},
        {"id": "r12pro", "brand": "Oppo", "colour": "Space Brown", "storage": "512",
         "model": "Oppo Reno 12 PRO 5G 512GB 12GB Space Brown", "model_number": "CPH2629"},
        {"id": "r13", "brand": "Oppo", "colour": "Luminous Blue", "storage": "512",
         "model": "Oppo Reno 13 5G Luminous Blue 12GB 512", "model_number": "CPH2689"},
        {"id": "a5promb", "brand": "Oppo", "colour": "Mocha Brown", "storage": "256",
         "model": "Oppo A5 Pro Mocha Brown 8GB RAM 256", "model_number": "CPH2695"},
        {"id": "a5proog", "brand": "Oppo", "colour": "Olive Green", "storage": "256",
         "model": "OPPO A5 Pro Olive Green 8GB RAM 256", "model_number": "CPH2695"},
        {"id": "a5mp", "brand": "Oppo", "colour": "Midnight Purple", "storage": "256",
         "model": "OPPO A5 Midnight Purple 8GB RAM 256", "model_number": "CPH2727"},
        # two genuinely indistinguishable rows (same model, colour, storage) -> must abstain
        {"id": "h72a", "brand": "Hisense", "colour": "Blue", "storage": "128",
         "model": "Hisense H72 Blue", "model_number": "HLTE270E"},
        {"id": "h72b", "brand": "Hisense", "colour": "Blue", "storage": "128",
         "model": "Hisense H72 Blue", "model_number": "HLTE270E"},
        {"id": "v30e", "brand": "Vivo", "colour": "Green", "storage": "256",
         "model": "vivo V30e Green 12GB RAM 256", "barcode": "6935117884578", "model_number": "V2339"},
    ]
    res = {r["source_id"]: r for r in match.match(S, T)}
    want = {  # source -> expected SKU, or None for must-abstain
        "a40sb": "OA40SBB", "a40sp": "OA40SBL",        # colour variants split, not collapsed
        "r12fog": "OR12FOG", "r12fmg": "OR12FMG",      # near-dup of reno13, blocked apart
        "r12pro": "CPH2629",                            # model# in SKU (strong key)
        "r13": "OR13BL",                                # not cross-assigned to reno12f
        "a5proog": "AOPROG", "a5mp": "OA5MP",          # a5 vs a5pro blocked apart
        "a5promb": None,                                # no listed SKU -> abstain, not a wrong guess
        "h72a": None, "h72b": None,                     # same colour+storage -> indistinguishable
        "v30e": "6935117884578",                        # barcode strong key
    }
    for sid, exp in want.items():
        got = res[sid]
        if exp is None:
            check(f"match: {sid} abstains", got["status"] == "abstained", str(got))
        else:
            check(f"match: {sid} -> {exp}", got["status"] == "matched" and got["sku"] == exp, str(got))
    skus = [r["sku"] for r in res.values() if r["status"] == "matched"]
    check("match: no SKU reused across rows", len(skus) == len(set(skus)), str(skus))


# --- semantic field mapping (W1: alias map kills duplicate synonym columns) --
def test_fields_semantic():
    import fields
    tab = {"roles": {"spec_start": 3, "spec_end": 5, "header_row": 1},
           "labels": {"3": "Tilt function:", "4": "Chair upholstery:", "5": "Gas lift safety class:"}}
    aliases = {"Tilt function": ["tilt mechanism"], "Chair upholstery": ["material"],
               "Gas lift safety class": ["gas lift"], "Category": ["product type"]}
    dec = {x["field"]: x for x in fields.reconcile(
        tab, [{"field": "Tilt mechanism"}, {"field": "Material"}, {"field": "Gas lift"},
              {"field": "Product type"}, {"field": "Whatsit rating"}],
        aliases=aliases, required_fields=["Category", "SKU", "EAN", "TSIN"])}
    check("fields/alias: 'Tilt mechanism' -> existing 'Tilt function:' column (no duplicate)",
          dec["Tilt mechanism"]["action"] == "existing" and dec["Tilt mechanism"]["col"] == 3)
    check("fields/alias: 'Material' -> 'Chair upholstery:'",
          dec["Material"]["action"] == "existing" and dec["Material"]["col"] == 4)
    check("fields/alias: 'Gas lift' -> 'Gas lift safety class:'",
          dec["Gas lift"]["action"] == "existing" and dec["Gas lift"]["col"] == 5)
    check("fields/alias: 'Product type' dropped as synonym of client-owned Category",
          dec["Product type"]["action"] == "drop")
    check("fields/alias: a genuinely new field is flagged uncertain, not silently appended",
          dec["Whatsit rating"]["action"] == "new" and dec["Whatsit rating"].get("uncertain") is True)


# --- three-state answer model (W0: value / sourced-"No" / unknown) ----------
def test_answer_kind():
    import confidence as C
    # a sourced NEGATIVE: feature provably absent from the complete table -> "No"
    absent = {"tab": "T", "row": 2, "field": "headrest", "column": 5, "value": "No",
              "provenance": "manufacturer", "answer_kind": "absent",
              "source_url": "https://maker/meta-bk",
              "note": "not present in the complete published spec table"}
    check("answer_kind: absent validates clean (note + url, no snippet needed)",
          ledger.validate([absent]) == [], str(ledger.validate([absent])))
    check("answer_kind: absent scored high and eligible",
          C.score_entry(absent)[0] == 88 and C.eligible(absent))
    bad = {"tab": "T", "row": 3, "field": "lumbar", "column": 6, "value": "No",
           "provenance": "manufacturer", "answer_kind": "absent"}
    check("answer_kind: absent without source_url/note is rejected",
          any("absent" in v for v in ledger.validate([bad])))
    unsure = dict(absent, completeness_uncertain=True)
    check("answer_kind: absent under uncertain completeness scored low and held",
          C.score_entry(unsure)[0] == 50 and not C.eligible(unsure))
    # gap taxonomy in needs_you: web-targetable blank is chased; structural is not
    import runner
    wb = Workbook(); ws = wb.active; ws.title = "Chairs"
    for c, h in enumerate(["Title", "Packaging weight:", "TSIN:"], 1):
        ws.cell(1, c, h)
    ws.cell(2, 1, "Chair A")  # both spec cells blank
    os.makedirs("/tmp/th", exist_ok=True); wb.save("/tmp/th/chairs.xlsx")
    schema = {"tabs": {"Chairs": {"roles": {"title": 1, "spec_start": 2, "spec_end": 2, "tsin": 3},
                                  "labels": {}, "header_row": 1, "first_data_row": 2, "example_rows": [],
                                  "gap_classes": {"structural_blank": ["TSIN"]}}}}
    gapline = " ".join(n for n in runner.needs_you(schema, [], "/tmp/th/chairs.xlsx") if n.startswith("gap "))
    check("answer_kind: web-targetable blank is chased",
          "web_targetable" in gapline and "Packaging weight" in gapline, gapline)
    check("answer_kind: structural blank (TSIN) is not nagged", "TSIN" not in gapline, gapline)


# --- delivery copy (W2) and tranche consolidation (W7) ----------------------
def test_delivery():
    import delivery
    from openpyxl import load_workbook
    wb = Workbook(); ws = wb.active; ws.title = "Chairs"
    for c, h in enumerate(["Title", "What's in the Box", "Seat depth", "Confidence",
                           "Check notes", "Source (root)", "Operating temperature"], 1):
        ws.cell(1, c, h)
    ws.cell(2, 1, "Chair"); ws.cell(2, 2, "1 x Chair"); ws.cell(2, 3, "520mm")
    ws.cell(2, 4, "93"); ws.cell(2, 5, "ok"); ws.cell(2, 6, "http://x")  # col 7 blank+hidden
    ws.column_dimensions["G"].hidden = True
    wb.create_sheet("Confidence key").cell(1, 1, "key")
    os.makedirs("/tmp/th", exist_ok=True); src = "/tmp/th/an.xlsx"; wb.save(src)
    w = load_workbook(delivery.make_delivery(src, "/tmp/th/del.xlsx")); s = w["Chairs"]
    hdrs = [s.cell(1, c).value for c in range(1, s.max_column + 1)]
    check("delivery: analysis columns stripped",
          not any(str(h).lower().startswith(("confidence", "check", "source")) for h in hdrs), str(hdrs))
    check("delivery: hidden blank column dropped", "Operating temperature" not in hdrs, str(hdrs))
    check("delivery: confidence-key sheet removed", "Confidence key" not in w.sheetnames)
    check("delivery: value preserved and formatted",
          s.cell(2, 3).value == "520mm" and s.row_dimensions[1].height == 30)


def test_consolidate():
    import consolidate
    from openpyxl import load_workbook
    wb1 = Workbook(); ws1 = wb1.active; ws1.title = "Keyboards"
    ws1.append(["Title", "Switch type", "SKU"]); ws1.append(["KB One", "Red", "K1"])
    os.makedirs("/tmp/th", exist_ok=True); p1 = "/tmp/th/t1.xlsx"; wb1.save(p1)
    wb2 = Workbook(); ws2 = wb2.active; ws2.title = "Keyboards (again)"
    ws2.append(["Title", "Switch type", "Hot-swap", "SKU"]); ws2.append(["KB Two", "Blue", "Yes", "K2"])
    p2 = "/tmp/th/t2.xlsx"; wb2.save(p2)
    w = load_workbook(consolidate.consolidate([p1, p2], "/tmp/th/cons.xlsx"))
    check("consolidate: one tab per category (no 'again')", w.sheetnames == ["Keyboards"], str(w.sheetnames))
    s = w["Keyboards"]; hdrs = [s.cell(1, c).value for c in range(1, s.max_column + 1)]
    check("consolidate: columns unioned (Hot-swap kept)", "Hot-swap" in hdrs, str(hdrs))
    titles = [s.cell(r, 1).value for r in range(2, s.max_row + 1)]
    check("consolidate: both rows present, none lost", titles == ["KB One", "KB Two"], str(titles))
    hi = hdrs.index("Hot-swap") + 1
    check("consolidate: a column a tranche lacks aligns blank, not shifted",
          s.cell(2, hi).value in (None, ""))


# --- completeness lone-product blind spot + EAN integrity (W9) ---------------
def test_completeness_lone():
    import completeness
    wb = Workbook(); ws = wb.active; ws.title = "Chairs"
    for c, h in enumerate(["Title", "Backrest height", "Seat depth", "Brand:", "SKU:"], 1):
        ws.cell(1, c, h)
    ws.append(["Chair A", "685mm", "520mm", "Endorfy", "EY8A005"])  # lone product, no source_field_count
    p = os.path.join(HERE, "_tmp_lone.xlsx"); wb.save(p)
    f = completeness.check(p, {"tabs": {"Chairs": {}}}, ratio=0.6, floor=2)
    check("completeness: lone product without source_field_count flagged unverifiable",
          any(x["row"] is None and "unverifiable" in x["why"] for x in f), str(f))
    os.remove(p)


def test_column_coverage():
    import completeness
    wb = Workbook(); ws = wb.active; ws.title = "Specs"
    for c, h in enumerate(["Model", "Processor", "Colour depth", "SKU", "TSIN"], 1):
        ws.cell(1, c, h)
    ws.append(["P1", "SD1", "", "K1", ""])
    ws.append(["P2", "SD2", "", "K2", ""])
    ws.append(["P3", "SD3", "", "K3", ""])
    ws.append(["P4", "SD4", "", "", ""])   # SKU blank here (the coverage hole)
    path = os.path.join(HERE, "_tmp_cov.xlsx"); wb.save(path)
    schema = {"tabs": {"Specs": {"roles": {"title": 1}, "header_row": 1, "first_data_row": 2,
                                 "example_rows": [], "gap_classes": {"structural_blank": ["TSIN"]}}}}
    f = completeness.column_coverage(path, schema, mostly=0.5)
    kinds = {(x["kind"], x["col"]) for x in f}
    check("coverage: all-empty non-structural column flagged (the colour-depth class)",
          ("empty-column", "Colour depth") in kinds, str(kinds))
    check("coverage: a column its siblings fill, blank on one row, flagged (the Xiaomi-block class)",
          ("coverage-hole", "SKU") in kinds, str(kinds))
    check("coverage: structural-blank column (TSIN) not flagged",
          not any(col == "TSIN" for _, col in kinds))
    check("coverage: a fully-filled column is not flagged",
          not any(col in ("Model", "Processor") for _, col in kinds))
    os.remove(path)


def test_receipt_closure():
    # the symmetric twin of "no value without a source": a negative needs a receipt
    bad = {"tab": "Specs", "row": 2, "field": "SKU", "column": 3, "answer_kind": "deferred"}
    check("receipt: a 'deferred' with no search_receipt is rejected",
          any("search_receipt" in v for v in ledger.validate([bad])))
    good = {"tab": "Specs", "row": 2, "field": "SKU", "column": 3, "answer_kind": "deferred",
            "search_receipt": "searched mi.com, gsmarena, takealot - SKU not public; distributor master"}
    check("receipt: a 'deferred' WITH a search_receipt validates", ledger.validate([good]) == [], str(ledger.validate([good])))
    import completeness
    wb = Workbook(); ws = wb.active; ws.title = "Specs"
    for c, h in enumerate(["Model", "Processor", "SKU", "TSIN"], 1):
        ws.cell(1, c, h)
    ws.append(["P1", "SD", "", ""])   # SKU blank (must be closed), TSIN blank (structural, exempt)
    os.makedirs("/tmp/th", exist_ok=True); p = "/tmp/th/closure.xlsx"; wb.save(p)
    schema = {"tabs": {"Specs": {"roles": {"title": 1}, "header_row": 1, "first_data_row": 2,
                                 "example_rows": [], "gap_classes": {"structural_blank": ["TSIN"]}}}}
    cols = {x["col"] for x in completeness.coverage_closure(p, schema, [])}
    check("receipt: a blank with no closing record is unresolved", "SKU" in cols, str(cols))
    check("receipt: a structural blank (TSIN) is exempt from closure", "TSIN" not in cols)
    closed = {x["col"] for x in completeness.coverage_closure(p, schema, [good])}
    check("receipt: a receipted 'deferred' closes the blank (no longer unresolved)", "SKU" not in closed, str(closed))
    os.remove(p)


def test_closure_reflex():
    import closure, completeness, delivery, ledger
    from openpyxl import load_workbook
    wb = Workbook(); ws = wb.active; ws.title = "Specs"
    for c, h in enumerate(["Model", "Model number", "Processor", "SKU", "TSIN"], 1):
        ws.cell(1, c, h)
    ws.append(["P1", "M-1", "", "", ""])   # Processor blank (spec->auto), SKU blank (identity->defer), TSIN structural
    os.makedirs("/tmp/th", exist_ok=True); p = "/tmp/th/reflex.xlsx"; wb.save(p)
    schema = {"tabs": {"Specs": {"roles": {"title": 1}, "header_row": 1, "first_data_row": 2,
                                 "example_rows": [], "gap_classes": {"structural_blank": ["TSIN"]}}}}
    findings = completeness.coverage_closure(p, schema, [])
    cols = {x["col"] for x in findings}
    check("reflex/gate: Processor & SKU unresolved, TSIN exempt",
          {"Processor", "SKU"} <= cols and "TSIN" not in cols, str(cols))
    briefs = closure.sourcing_brief(p, schema, findings)
    b = briefs[("Specs", 2)]
    check("reflex/brief: spec field is auto-sourced", "Processor" in b["auto_source"])
    check("reflex/brief: identity (SKU) is DEFERRED, never web-sourced (human-in-the-loop)",
          any(f == "SKU" for f, _ in b["defer"]) and "SKU" not in b["auto_source"])
    check("reflex/brief: product identity captured for the search", b["identity"].get("Model number") == "M-1")
    n = closure.merge_sourced(p, {"M-1": {"Processor": {"value": "Helio G99"}}}, schema)
    check("reflex/merge: spec value filled by model number (variant-safe)",
          n == 1 and load_workbook(p)["Specs"].cell(2, 3).value == "Helio G99")
    findings2 = completeness.coverage_closure(p, schema, [])
    receipts = {("Specs", 2, f["column"]): "no master; vendor confirm" for f in findings2}
    entries, unreceipted = closure.receipt_residue(findings2, receipts)
    check("reflex/receipt: residue closed (none unreceipted)", not unreceipted, str(unreceipted))
    check("reflex/receipt: deferred entries validate", ledger.validate(entries) == [], str(ledger.validate(entries)))
    try:
        delivery.make_delivery(p, "/tmp/th/reflex_del.xlsx", schema=schema, entries=[]); refused = False
    except ValueError:
        refused = True
    check("reflex/delivery: REFUSES to ship an unresolved blank", refused)
    delivery.make_delivery(p, "/tmp/th/reflex_del.xlsx", schema=schema, entries=entries)
    check("reflex/delivery: ships once the blank is receipted as deferred",
          os.path.exists("/tmp/th/reflex_del.xlsx"))
    os.remove(p)


def test_ean_integrity():
    import audit
    wb = Workbook(); ws = wb.active; ws.title = "Mice"
    for c, h in {1: "Title", 2: "Spec", 3: "SKU", 4: "EAN"}.items():
        ws.cell(1, c, h)
    ws.cell(2, 1, "A"); ws.cell(2, 3, "SKU-A"); ws.cell(2, 4, "5903018666495")   # valid EAN-13
    ws.cell(3, 1, "B"); ws.cell(3, 3, "SKU-B"); ws.cell(3, 4, "5903018666495")   # same EAN, different SKU
    ws.cell(4, 1, "C"); ws.cell(4, 3, "SKU-C"); ws.cell(4, 4, "1234567890123")   # bad check digit
    os.makedirs("/tmp/th", exist_ok=True); wb.save("/tmp/th/ean.xlsx")
    schema = {"tabs": {"Mice": {"roles": {"title": 1, "spec_start": 2, "spec_end": 2, "sku": 3, "ean": 4},
                                "labels": {}, "header_row": 1, "first_data_row": 2, "example_rows": []}}}
    ledger.save("/tmp/th/led.json", [])
    F = audit.audit("/tmp/th/ean.xlsx", schema, "/tmp/th/led.json", "/tmp/th/rep.md", profiles={})
    msgs = " ".join(m for *_, m in F)
    check("ean: shared barcode across different SKUs flagged", "shared across different SKUs" in msgs, msgs[:160])
    check("ean: invalid check digit flagged", "not a valid barcode" in msgs)
    check("ean: a valid EAN-13 is accepted", "5903018666495' is not a valid" not in msgs)


# --- Excel formatting pass (W6) ---------------------------------------------
def test_format_xlsx():
    import format_xlsx as FX
    from openpyxl import load_workbook
    wb = Workbook(); ws = wb.active; ws.title = "Chairs"
    ws.cell(1, 1, "Title"); ws.cell(1, 2, "What's in the Box"); ws.cell(1, 3, "Seat depth")
    ws.cell(2, 1, "Chair"); ws.cell(2, 2, "1 x Chair"); ws.cell(2, 3, "520mm")
    os.makedirs("/tmp/th", exist_ok=True); p = "/tmp/th/fmt.xlsx"; wb.save(p)
    FX.format_workbook(p)
    s = load_workbook(p)["Chairs"]
    check("format: row height 30",
          s.row_dimensions[1].height == 30 and s.row_dimensions[2].height == 30)
    check("format: A & B = 45, others = 16",
          round(s.column_dimensions["A"].width) == 45 and round(s.column_dimensions["B"].width) == 45
          and round(s.column_dimensions["C"].width) == 16)
    check("format: header centred+bold, body left, all wrap+middle",
          s.cell(1, 1).alignment.horizontal == "center" and s.cell(1, 1).font.bold
          and s.cell(2, 1).alignment.horizontal == "left"
          and s.cell(2, 1).alignment.vertical == "center" and s.cell(2, 1).alignment.wrap_text)
    check("format: thin borders across the data range",
          s.cell(2, 3).border.left.style == "thin" and s.cell(2, 3).border.bottom.style == "thin")


# --- derivation with provenance (W3) ----------------------------------------
def test_derive():
    import derive, ledger
    import confidence as C
    c = derive.colour_from_title("Oppo A40 (Sparkle Black, 128 GB)")
    check("derive: colour pulled from title", c is not None and "black" in c[0].lower(), str(c))
    check("derive: no colour word in title -> abstain",
          derive.colour_from_title("Endorfy Meta BK Ergonomic Gaming Chair") is None)
    h = derive.height_from_dimensions("1300 x 690 x 640mm", "HWL")
    check("derive: total height = published H axis (not guessed)", h is not None and h[0] == "1300mm", str(h))
    check("derive: refuses to pick a height without a published axis order",
          derive.height_from_dimensions("1300 x 690 x 640mm", "") is None)
    e = derive.make_entry("Chairs", 2, 30, "Total height", "1300mm", h[1])
    check("derive: derived entry validates (provenance + note)", ledger.validate([e]) == [], str(ledger.validate([e])))
    check("derive: derived entry scored 70 and eligible", C.score_entry(e)[0] == 70 and C.eligible(e))


# --- house style guide + charset (W5) ---------------------------------------
def test_standardise_style():
    import standardise as S
    f = lambda h, v: S.standardise_value(h, v)[0]
    check("style: weight no space", f("Product weight:", "23.3 kg") == "23.3kg")
    check("style: lone dimension no space", f("Seat depth:", "520 mm") == "520mm")
    check("style: dimension triple, x spaced, unit attached to last",
          f("Product dimensions:", "1300 × 690 × 640 mm") == "1300 x 690 x 640mm")
    check("style: cm triple scaled to mm",
          f("Packaging dimensions:", "76 x 63.5 x 37.5 cm") == "760 x 635 x 375mm")
    check("style: angle range spelled out, + dropped",
          f("Backrest incline:", "100° to 125°") == "100 to 125 degrees")
    check("style: signed angle range, + dropped, - kept",
          f("Max seat tilting angle:", "-90° to +90°") == "-90 to 90 degrees")
    check("style: temperature spelled out with Celsius",
          f("Operating temperature:", "-10 °C / 40 °C") == "-10 to 40 degrees Celsius")
    check("style: multiplication sign becomes x",
          "×" not in f("Wheels:", "2 × 60 mm caster"))
    check("style: smart quotes straightened",
          f("Bearing type:", "the “big” one") == 'the "big" one')
    check("style: prose quantity spacing 1x -> 1 x",
          f("What's in the Box:", "1x Chair, 1x Assembly kit") == "1 x Chair, 1 x Assembly kit")
    check("style: prose keeps pack dimensions intact (no numeric reformat)",
          "120mm" in f("What's in the Box:", "2x 120mm fan"))
    check("style: network '4G'/'5G' not mangled into grams",
          f("Network:", "4G") == "4G" and f("Network:", "4G LTE") == "4G LTE")
    check("style: real weight still normalises (lowercase g)", f("Weight:", "200 g") == "200g")
    check("style: warranty month spacing/casing", f("Warranty:", "12months") == "12 months"
          and f("Warranty:", "12 Months") == "12 months")
    # socket-style noise is FLAGGED, not auto-changed
    wb = Workbook(); ws = wb.active; ws.title = "Coolers"
    for c, hh in enumerate(["Title", "Socket compatibility:", "SKU:"], 1):
        ws.cell(1, c, hh)
    ws.cell(2, 1, "Cooler"); ws.cell(2, 2, "AM3(+), AM2(+)"); ws.cell(2, 3, "X1")
    os.makedirs("/tmp/th", exist_ok=True); p = "/tmp/th/style.xlsx"; wb.save(p)
    fixes, flags = S.check(p)
    check("style: socket-noise '(+)' flagged, not auto-stripped",
          any("socket" in fl.get("why", "") for fl in flags)
          and not any(fx["col"].strip() == "Socket compatibility:" for fx in fixes))


def main():
    for fn in [test_engine_fixtures, test_engine_guards, test_conflicts, test_ledger_validate,
               test_profile_builder, test_linkcheck_offline, test_duplicate_sku,
               test_jobspec, test_schema_store, test_runner, test_primary_source,
               test_confidence, test_fields, test_hide_blank_columns,
               test_from_spec_table, test_completeness, test_match, test_answer_kind,
               test_fields_semantic, test_standardise_style, test_format_xlsx, test_derive,
               test_completeness_lone, test_ean_integrity, test_delivery, test_consolidate,
               test_column_coverage, test_receipt_closure, test_closure_reflex]:
        try:
            fn()
        except Exception as ex:  # noqa
            FAIL.append(f"{fn.__name__} raised {type(ex).__name__}: {ex}")
    print(f"\n{len(PASS)} passed, {len(FAIL)} failed")
    for p in PASS:
        print("  PASS", p)
    for f in FAIL:
        print("  FAIL", f)
    sys.exit(1 if FAIL else 0)


if __name__ == "__main__":
    main()
