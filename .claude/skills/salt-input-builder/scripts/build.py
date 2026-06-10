"""Deterministic build. Writes the workbook ONLY from a validated ledger.

It introduces no new facts: every cell value is copied from a ledger entry,
which itself carries a snippet it was copied from. It refuses to run on a
ledger with validation violations. It colours each row's Confidence cell by
the row's most conservative provenance tier, preserves the CONCAT/listing
column, and writes a liveness-checked source link per row.
"""
from __future__ import annotations
import sys, copy
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import common, ledger as L, linkcheck, confidence
from common import DEFAULT_COLOURS, BODY_FONT


def _hide_blank_columns(ws, t):
    """Hide spec/extra columns that are blank across every product row; un-hide
    any that have data. Reversible, index-stable, never deletes."""
    roles = t["roles"]
    hr = t.get("header_row", 1)
    title_col = roles.get("title", 1)
    data_rows = [r for r in range(hr + 1, ws.max_row + 1) if ws.cell(r, title_col).value not in (None, "")]
    if not data_rows:
        return 0
    a, b = roles.get("spec_start"), roles.get("spec_end")
    cols = set(range(a, b + 1)) if a and b else set()
    cols |= set((t.get("extra_fields") or {}).values())
    hidden = 0
    for c in cols:
        has_data = any(ws.cell(r, c).value not in (None, "") for r in data_rows)
        ws.column_dimensions[get_column_letter(c)].hidden = not has_data
        if not has_data:
            hidden += 1
    return hidden


def _write_key_sheet(wb, colours):
    """A readable key: colour code, score bands, and how each score is derived."""
    name = "Confidence key"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    bold = Font(bold=True, **BODY_FONT)
    ws.cell(1, 1, value="Confidence key - how each score is derived").font = Font(bold=True, size=12, name="Arial")
    bands = [("85-100", "High", "manufacturer", "Confirmed, or client identity", "Manufacturer source, or a client-provided identity field"),
             ("60-84", "Medium", "retailer", "Corroborated, no manufacturer", "Two or more independent sources agree (70 = 2 sources, 82 = 3+); spot-check"),
             ("0-59", "Low", "unverifiable", "Needs verification before use", "Single uncorroborated source (40), or sources disagree (35)")]
    ws.cell(3, 1, value="Score").font = bold
    ws.cell(3, 2, value="Band").font = bold
    ws.cell(3, 3, value="Colour").font = bold
    ws.cell(3, 4, value="Meaning").font = bold
    ws.cell(3, 5, value="How it is derived").font = bold
    for i, (rng, lab, key, meaning, how) in enumerate(bands, start=4):
        ws.cell(i, 1, value=rng).font = Font(**BODY_FONT)
        ws.cell(i, 2, value=lab).font = Font(**BODY_FONT)
        common.set_fill(ws, i, 3, colours[key]["fill"])
        ws.cell(i, 3).font = Font(color="FF" + colours[key]["font"], **BODY_FONT)
        ws.cell(i, 4, value=meaning).font = Font(**BODY_FONT)
        ws.cell(i, 5, value=how).font = Font(**BODY_FONT)
    derivation = [
        "Exact scores: manufacturer source = 90; manufacturer + a second source = 97; client identity = 85;",
        "3+ independent sources agree = 82; 2 independent sources agree = 70; derived/computed = 70;",
        "single uncorroborated source = 40; sources disagree = 35.",
        "",
        "Each data cell is shaded by its own band. The row Confidence is the mean of that row's cells.",
        "The Check notes column lists every cell scoring under 85, with its value and reason.",
        "The full per-cell map (value, score, reason, source count) is exported to confidence_map.json for SALT.",
    ]
    for j, line in enumerate(derivation, start=8):
        ws.cell(j, 1, value=line).font = Font(**BODY_FONT)
    for c, w in ((1, 12), (2, 10), (3, 10), (4, 30), (5, 70)):
        ws.column_dimensions[ws.cell(1, c).column_letter].width = w


def build(template, schema, ledger_path, out_path, colours=None, check_links=True):
    colours = {**DEFAULT_COLOURS, **(colours or {})}
    entries = L.load(ledger_path)

    violations = L.validate(entries)
    if violations:
        print("LEDGER VALIDATION FAILED - build refused. Fix these first:")
        for v in violations:
            print("  -", v)
        return {"ok": False, "violations": violations}

    # liveness-check every distinct source url
    link_status = {}
    if check_links:
        urls = [e["source_url"] for e in entries if e.get("source_url")]
        link_status = linkcheck.check_many(urls)
        for e in entries:
            u = e.get("source_url")
            if u:
                ok, _ = link_status[u]
                e["link_ok"] = ok

    wb = load_workbook(template)
    needs_you = []
    written = 0
    confidence.annotate(entries)  # stamp confidence, reason, eligible, band
    band_key = {"high": "manufacturer", "medium": "retailer", "low": "unverifiable"}

    for e in entries:
        ws = wb[e["tab"]]
        col = e["column"]
        # never clobber the CONCAT/listing-output column
        concat_col = schema["tabs"][e["tab"]]["roles"].get("concat")
        if col == concat_col:
            continue
        if e.get("answer_kind") in ("deferred", "unknown"):
            continue  # a record of a failed search, not a value - the cell stays blank
        # write every sourced value; weak ones are flagged, not blanked
        common.write_cell(ws, e["row"], col, e["value"])
        common.set_fill(ws, e["row"], col, colours[band_key[e["band"]]]["fill"])
        written += 1
        if not e.get("eligible"):
            needs_you.append(f"{e['tab']} r{e['row']}: '{e.get('field')}' written but FLAGGED - {e['confidence_reason']}")

    # per-row: mean confidence into the row, detailed check notes, source link
    rows = sorted({(e["tab"], e["row"]) for e in entries})
    for tab, row in rows:
        t = schema["tabs"][tab]
        roles = t["roles"]
        ws = wb[tab]
        hr = t.get("header_row", 1)
        sourced = [e for e in entries if e["tab"] == tab and e["row"] == row
                   and e["column"] != roles.get("concat")]
        if "confidence" in roles and sourced:
            mean = round(sum(e["confidence"] for e in sourced) / len(sourced))
            bk = band_key[confidence.band(mean)]
            cc = ws.cell(row, roles["confidence"], value=mean)
            common.set_fill(ws, row, roles["confidence"], colours[bk]["fill"])
            cc.font = Font(color="FF" + colours[bk]["font"], **BODY_FONT)
        if "notes" in roles and sourced:
            fnotes = t.get("field_notes", {})
            flags = []
            for e in sorted(sourced, key=lambda x: x["confidence"]):
                if e["confidence"] < 85:
                    lab = ws.cell(hr, e["column"]).value or e.get("field")
                    flags.append(f"{lab} = {e['value']}: {e['confidence_reason']}")
            # non-identical header-mapping disclosures (separate from value doubt)
            for e in sourced:
                fn = fnotes.get(str(e["column"]))
                if fn:
                    flags.append(f"MAPPING: {fn}")
            note = "; ".join(flags) if flags else "all populated values confirmed (manufacturer source or client identity)"
            nc = ws.cell(row, roles["notes"], value=note)
            nc.font = Font(**BODY_FONT)
            nc.alignment = common.Alignment(**common.TOP_WRAP)
        # source link
        url = L.primary_source(entries, tab, row)
        if "source" in roles:
            scol = roles["source"]
            if url:
                ok = link_status.get(url, (None, ""))[0] if check_links else None
                if ok is False:
                    needs_you.append(f"{tab} r{row}: source link appears DEAD ({url}) - replace before shipping")
                elif ok is None and check_links:
                    needs_you.append(f"{tab} r{row}: source link liveness UNVERIFIED ({url})")
                c = ws.cell(row, scol, value=url)
                c.hyperlink = url
                c.font = Font(name="Arial", size=10, color="FF0563C1", underline="single")
                c.alignment = common.Alignment(**common.TOP_WRAP)
            else:
                needs_you.append(f"{tab} r{row}: no source URL on any field for this row")

    # source conflicts among candidate observations (HARD ones must be resolved)
    for tab, row, col, field, sev, msg in L.detect_conflicts(entries):
        if sev == "HARD":
            needs_you.append(f"{tab} r{row}: source conflict on {field} - {msg}")

    # headers for the columns we populate, if blank
    for tab, t in schema["tabs"].items():
        ws = wb[tab]; hr = t.get("header_row", 1); roles = t["roles"]
        for role, label in (("source", "Source (root)"), ("confidence", "Confidence (sanity check)"),
                            ("notes", "Check notes (flagged items)")):
            col = roles.get(role)
            if col and ws.cell(hr, col).value in (None, ""):
                ws.cell(hr, col, value=label).font = Font(bold=True, **BODY_FONT)

    # hide columns that are entirely blank across this tab's products (idempotent:
    # a column that gains data on a later run is un-hidden). Never deletes.
    hidden_total = 0
    for tab, t in schema["tabs"].items():
        hidden_total += _hide_blank_columns(wb[tab], t)

    _write_key_sheet(wb, colours)
    wb.save(out_path)
    L.save(ledger_path, entries)  # persist link_ok back into the ledger
    print(f"BUILD OK: {written} cells written from ledger across {len(rows)} rows -> {out_path}")
    print(f"hid {hidden_total} fully-blank spec columns across {len(schema['tabs'])} tabs")
    if needs_you:
        print("NEEDS YOU:")
        for n in needs_you:
            print("  -", n)
    return {"ok": True, "written": written, "rows": len(rows), "needs_you": needs_you}


if __name__ == "__main__":
    # build.py <template.xlsx> <schema.json> <ledger.json> <out.xlsx>
    import common as _c
    build(sys.argv[1], _c.load_json(sys.argv[2]), sys.argv[3], sys.argv[4])
