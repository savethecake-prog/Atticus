"""House-style standardiser: enforce one consistent output format.

completeness.py guards against thin data. This guards the orthogonal failure:
the same fact written five different ways. SALT reproduces the specifications
table verbatim (specs in == specs out), so any inconsistency on the input sheet
lands directly on the live customer listing, and two products on the same tab
can show dimensions in two different styles purely because they were typed
differently.

Division of labour:
  * AUTO-FIX (safe, mechanical, value-preserving): dimension format and unit,
    weight format, temperature format, Yes/No casing, UK spelling, brand casing.
  * FLAG ONLY (a machine must not guess these): a value that looks like it is in
    the wrong field (e.g. a recline angle inside a seat-tilt cell), an identical
    long value duplicated across two different fields, layout deviations
    (column A name, analysis columns still present).

Auto-fix is value-preserving: cm and m are converted to mm by scaling the
numbers, decimal commas become points, and nothing changes the physical meaning.
Structural conversion (renaming headers, adding the required tail, reordering
columns, converting a tranche layout to the house layout) is NOT done here; that
belongs to the build/correction pass. This module only standardises and reports.
"""
from __future__ import annotations
import re
from openpyxl import load_workbook

# ---- canonical configuration -------------------------------------------------

# columns whose VALUES must never be reformatted (identifiers and plumbing).
# NB: "image"/"video" are NOT bare substrings here — they collided with spec
# fields like "Video memory capacity" / "Video base clock", which were then
# silently skipped from compaction. Media columns are matched by _is_media().
SKIP_VALUE = ("sku", "ean", "tsin", "barcode", "model number", "model code",
              "product code", "source", "confidence", "check", "status",
              "pl2", "rrp", "qty", "title")

# media-link columns (their URLs must never be reformatted), matched precisely so
# a spec field that merely starts with "Video"/"Image" is not caught.
_MEDIA_EXACT = {"image", "images", "video", "videos", "imageurl", "imageurls",
                "videourl", "videourls", "mainimage", "additionalimages",
                "media", "mediaurl"}
def _is_media(header):
    k = re.sub(r"[^a-z0-9]", "", header.lower())
    if k in _MEDIA_EXACT:
        return True
    return bool(re.search(r"image|video", header, re.I)
                and re.search(r"url|link|path|src", header, re.I))

# prose columns: only spelling is normalised, never numeric reformatting
PROSE = ("what", "in the box", "why", "description", "feature", "subtitle", "note")

# US -> UK, whole word, case of first letter preserved
UK_SPELL = {
    "color": "colour", "colors": "colours", "colored": "coloured",
    "fiber": "fibre", "fibers": "fibres",
    "aluminum": "aluminium", "gray": "grey",
    "customizable": "customisable", "organizer": "organiser",
    "fulfill": "fulfil", "center": "centre", "meter": "metre", "liter": "litre",
}
# controlled brand vocabulary -> exact casing
BRANDS = {
    "cougar": "Cougar", "thermalright": "Thermalright", "asus": "ASUS",
    "thermal grizzly": "Thermal Grizzly", "polartherm": "Polartherm",
    "endorfy": "Endorfy", "thermaltake": "Thermaltake",
}
# placement heuristics: (header substring, value-words that don't belong, why)
SUSPICIOUS = [
    ("seat tilt", ("recline", "backrest"),
     "a recline/backrest value sitting in a seat-tilt field"),
    ("backrest angle", ("seat tilt", "seat tilting"),
     "a seat-tilt value sitting in a backrest-angle field"),
]
# values short/common enough to legitimately repeat across fields (no dup flag)
_REPEATABLE = {"yes", "no", "n/a", "na", "none", "black", "white", "grey",
               "rgb", "argb", "standard", "1", "0",
               "not specified", "not applicable", "not published", "tbc"}

_DIM = re.compile(
    r"(?<![\d.])(\d+(?:[.,]\d+)?)\s*[x×]\s*(\d+(?:[.,]\d+)?)"
    r"(?:\s*[x×]\s*(\d+(?:[.,]\d+)?))?\s*(mm|cm|m)?\b", re.I)
_LEN_ONE = re.compile(r"(?<![\d.])(\d+(?:[.,]\d+)?)\s+(mm|cm|m)\b", re.I)  # lone "520 mm" -> "520mm"
# only the required Product/Packaging dimension fields scale cm/m to mm; every
# other dimension field keeps its native (often cm) unit and is compact-only.
_SCALE_DIM = re.compile(r"(product|packaging)\s*dimension", re.I)
# contrast ratio carries no thousands separator ("1,000:1" -> "1000:1"); other
# fields (e.g. storage speeds) keep their commas, so this is keyed to contrast.
_CONTRAST = re.compile(r"contrast", re.I)
_THOUSANDS = re.compile(r"(?<=\d),(?=\d)")
_WT = re.compile(r"(?<![\d.])(\d+(?:[.,]\d+)?)\s*([kK][gG]|g)\b")  # 'g' lowercase only: "4G"/"5G" network is not grams
# compact the space between a number and an attached unit, for machine-readable
# ingestion (Christopher's compact rule). A closed whitelist of units only, so
# natural-language ("8 megapixel", "Up to 4") is never touched. cd/m² last so the
# slash/superscript survive intact.
_UNIT_COMPACT = re.compile(
    r"(?<![A-Za-z0-9])(\d+(?:[.,]\d+)?)\s+("
    r"GHz|MHz|kHz|KHz|Hz|Gbps|Mbps|GB/s|MB/s|TB|GB|MB|KB|mAh|nm|ms|DPI|dpi|cd/m²|W"
    r")(?![A-Za-z])", re.I)
_MONTHS = re.compile(r"\b(\d+)\s*[Mm]onths\b")                     # "12months"/"12 Months" -> "12 months"
# temperatures carry a Celsius unit; angles carry a bare degree sign. Both spelled out.
_TEMP_RANGE = re.compile(r"([+-]?\d+)\s*°?\s*C\s*(?:to|/|–|—|-)\s*([+-]?\d+)\s*°?\s*C", re.I)
_TEMP_ONE = re.compile(r"([+-]?\d+)\s*°\s*C\b", re.I)
_ANG_RANGE = re.compile(r"([+-]?\d+)\s*°\s*(?:to|/|–|—|-)\s*([+-]?\d+)\s*°")
_ANG_ONE = re.compile(r"([+-]?\d+)\s*°(?!\s*C)")
_QTY = re.compile(r"\b(\d+)\s*[x×]\s*(?=[A-Za-z])")        # "1x Chair" / "1 x chair" -> "1 x "
_SOCKET = re.compile(r"[A-Za-z0-9]+\(\+\)")               # AM3(+) noise -> FLAG, never auto-strip

# ASCII-safe substitutions (the "stick to a safe charset, no stray symbols" rule).
# The degree sign is deliberately NOT dropped here: it is spelled out to
# "degrees"/"degrees Celsius" by the passes above so the unit survives as words.
_CHARSET = {
    " ": " ", " ": " ", " ": " ",      # non-breaking spaces
    "‑": "-", "‒": "-", "–": "-", "—": "-", "−": "-",  # hyphens/dashes/minus
    "‘": "'", "’": "'", "“": '"', "”": '"',  # smart quotes
    "×": "x", "™": "", "®": "", "©": "", "…": "...",  # x, tm, r, c, ellipsis
}


def _num(s):
    s = s.replace(",", ".")
    f = float(s)
    return int(f) if f == int(f) else f


def _scale(s, factor):
    v = _num(s) * factor
    return str(int(v) if v == int(v) else round(v, 3))


def _norm_dim(m, scale_mm=True):
    a, b, c, unit = m.group(1), m.group(2), m.group(3), (m.group(4) or "").lower()
    parts = [a, b] + ([c] if c else [])
    if unit:
        if scale_mm:
            # required Product/Packaging dimension fields: scale cm/m to mm
            factor = {"cm": 10, "m": 1000, "mm": 1}[unit]
            nums = [_scale(p, factor) for p in parts]
            return "x".join(nums[:-1] + [nums[-1] + "mm"])
        # every other dimension field (GPU length, monitor physical dims): keep the
        # field's native unit, only compact — never scale to mm (schema is cm-native).
        nums = [str(_num(p)) for p in parts]
        return "x".join(nums) + unit
    # no unit (e.g. a resolution "1920 x 1080"): compact the separators only,
    # never scale and never assert mm. The unitless-dimension flag still fires.
    return "x".join(str(_num(p)) for p in parts)


def _norm_len_one(m):
    return f"{_num(m.group(1))}{m.group(2).lower()}"   # lone "520 mm" -> "520mm" (no scaling)


def _round_half_up(x):
    import math
    return int(math.floor(x + 0.5))


def _norm_wt(m):
    """Compact, and apply the threshold rule: <1kg -> grams (no decimals,
    <0.5 down / >=0.5 up), >=1kg -> kg; >=1000g -> kg."""
    n, unit = _num(m.group(1)), m.group(2).lower()
    grams = n * 1000 if unit == "kg" else n
    if grams < 1000:
        return f"{_round_half_up(grams)}g"
    kg = grams / 1000
    return f"{_num(str(int(kg) if kg == int(kg) else round(kg, 3)))}kg"


def _ascii_safe(s):
    for k, v in _CHARSET.items():
        s = s.replace(k, v)
    return s


def _pm(n):
    return n[1:] if n.startswith("+") else n           # drop a leading +, keep a leading -


def _spell_angles_temps(s):
    s = _TEMP_RANGE.sub(lambda m: f"{_pm(m.group(1))} to {_pm(m.group(2))} degrees Celsius", s)
    s = _TEMP_ONE.sub(lambda m: f"{_pm(m.group(1))} degrees Celsius", s)
    s = _ANG_RANGE.sub(lambda m: f"{_pm(m.group(1))} to {_pm(m.group(2))} degrees", s)
    s = _ANG_ONE.sub(lambda m: f"{_pm(m.group(1))} degrees", s)
    s = s.replace("\u00b0C", " degrees Celsius").replace("\u00b0", " degrees")  # stray degree signs
    return re.sub(r"\s{2,}", " ", s)


def _uk(text):
    def repl(mt):
        w = mt.group(0)
        uk = UK_SPELL[w.lower()]
        return uk.capitalize() if w[0].isupper() else uk
    pat = re.compile(r"\b(" + "|".join(map(re.escape, UK_SPELL)) + r")\b", re.I)
    return pat.sub(repl, text)


def standardise_value(header, value):
    """Return (new_value, [fix notes]) for one cell. Empty notes means unchanged."""
    if value is None or not isinstance(value, str):
        return value, []
    h = header.lower()
    if _is_media(header) or any(k in h for k in SKIP_VALUE):
        return value, []
    # prose cells (What's in the Box, features, descriptions): charset + quantity
    # spacing + spelling only; never numeric reformatting ("2 x 120mm" is a quantity).
    if any(k in h for k in PROSE) or len(value) > 80:
        out, notes = value, []
        new = _ascii_safe(out)
        if new != out:
            notes.append("charset"); out = new
        new = _QTY.sub(lambda m: f"{m.group(1)} x ", out)
        if new != out:
            notes.append("quantity spacing"); out = new
        new = _uk(out)
        if new != out:
            notes.append("UK spelling"); out = new
        return out, notes
    out, notes = value, []
    scale_mm = bool(_SCALE_DIM.search(header))   # only required dimension fields scale to mm
    new = _ascii_safe(out)
    if new != out:
        notes.append("charset"); out = new
    new = _DIM.sub(lambda m: _norm_dim(m, scale_mm), out)
    if new != out:
        notes.append("dimension format/unit"); out = new
    new = _LEN_ONE.sub(_norm_len_one, out)
    if new != out:
        notes.append("unit spacing"); out = new
    if _CONTRAST.search(header):
        new = _THOUSANDS.sub("", out)
        if new != out:
            notes.append("contrast ratio commas"); out = new
    new = _UNIT_COMPACT.sub(lambda m: f"{m.group(1)}{m.group(2)}", out)
    if new != out:
        notes.append("measurement compacted"); out = new
    new = _WT.sub(_norm_wt, out)
    if new != out:
        notes.append("weight format"); out = new
    new = _MONTHS.sub(r"\1 months", out)
    if new != out:
        notes.append("month spacing"); out = new
    new = _spell_angles_temps(out)
    if new != out:
        notes.append("angle/temperature spelled out"); out = new
    if out.strip().lower() in ("yes", "no"):
        cap = out.strip().capitalize()
        if cap != out:
            notes.append("Yes/No casing"); out = cap
    new = _uk(out)
    if new != out:
        notes.append("UK spelling"); out = new
    return out, notes


def _headers(ws):
    return {c: str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)
            if ws.cell(1, c).value}


def check(path, fix_inplace=False):
    """Scan a workbook. Returns (fixes, flags).

    fixes: cells that differ from house style and can be safely auto-corrected.
    flags: deviations a machine must not change on its own.
    With fix_inplace=True the safe fixes are written into the loaded workbook and
    the modified workbook object is returned as a third element so the caller can
    save it.
    """
    wb = load_workbook(path)
    fixes, flags = [], []
    reported_pairs = set()   # (tab, {colA,colB}) so a redundant column pair flags once
    for ws in wb.worksheets:
        H = _headers(ws)
        a1 = str(ws.cell(1, 1).value or "")
        if a1 and a1.lower() != "title":
            flags.append({"tab": ws.title, "row": 1, "kind": "layout",
                          "why": f"column A is {a1!r}, house style is 'Title'"})
        for c, hdr in H.items():
            if any(k == hdr.strip().lower() for k in ("confidence", "check notes",
                                                      "source (root)", "source root", "source")):
                flags.append({"tab": ws.title, "row": 1, "kind": "layout",
                              "why": f"analysis column {hdr!r} should not ship in the upload copy"})
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value in (None, ""):
                continue
            seen = {}
            for c, hdr in H.items():
                v = ws.cell(r, c).value
                low_h = hdr.lower()
                # brand canonicalisation
                if low_h.strip().rstrip(":").strip() == "brand" and isinstance(v, str):
                    canon = BRANDS.get(v.strip().lower())
                    if canon and canon != v:
                        fixes.append({"tab": ws.title, "row": r, "col": hdr,
                                      "old": v, "new": canon, "why": "brand casing"})
                        if fix_inplace:
                            ws.cell(r, c).value = canon
                    continue
                # placement heuristics (flag only)
                if isinstance(v, str):
                    for hsub, bad, why in SUSPICIOUS:
                        if hsub in low_h and any(b in v.lower() for b in bad):
                            flags.append({"tab": ws.title, "row": r, "kind": "placement",
                                          "col": hdr, "value": v, "why": why})
                # duplicate long value across two different spec fields (flag only)
                if isinstance(v, str) and len(v.strip()) > 8 and v.strip().lower() not in _REPEATABLE \
                        and not any(k in low_h for k in SKIP_VALUE) and low_h.strip().rstrip(":").strip() != "brand":
                    if v.strip() in seen and seen[v.strip()] != hdr:
                        pair = (ws.title, frozenset((hdr, seen[v.strip()])))
                        if pair not in reported_pairs:
                            reported_pairs.add(pair)
                            flags.append({"tab": ws.title, "row": r, "kind": "placement",
                                          "col": hdr, "value": v,
                                          "why": f"identical value also in {seen[v.strip()]!r} - "
                                                 f"redundant columns or a placement error (recurs on other rows)"})
                    seen[v.strip()] = hdr
                # unitless dimension (flag only) - ONLY in genuine dimension/size
                # fields, so resolutions, slot counts and connector names that
                # happen to contain "x" are not mistaken for measurements.
                if isinstance(v, str) and ("dimension" in low_h or "size" in low_h) \
                        and re.search(r"\d\s*[x\u00d7]\s*\d", v) and not re.search(r"(mm|cm|\bm\b)", v.lower()):
                    flags.append({"tab": ws.title, "row": r, "kind": "format", "col": hdr,
                                  "value": v, "why": "dimension/size has no unit - cannot normalise to mm safely"})
                # socket-style noise like "AM3(+)" - flag for confirmation, never auto-strip
                if isinstance(v, str) and _SOCKET.search(v):
                    flags.append({"tab": ws.title, "row": r, "kind": "format", "col": hdr,
                                  "value": v, "why": "value carries socket-style noise like 'AM3(+)'; "
                                                     "confirm the intended plain form (e.g. 'AM3, AM2, & FM2')"})
                # safe value fixes
                new, notes = standardise_value(hdr, v)
                if notes:
                    fixes.append({"tab": ws.title, "row": r, "col": hdr,
                                  "old": v, "new": new, "why": ", ".join(notes)})
                    if fix_inplace:
                        ws.cell(r, c).value = new
    return (fixes, flags, wb) if fix_inplace else (fixes, flags)


def fix(path, out_path):
    fixes, flags, wb = check(path, fix_inplace=True)
    wb.save(out_path)
    return fixes, flags


def summarise(fixes, flags):
    byk = {}
    for f in fixes:
        byk[f["why"]] = byk.get(f["why"], 0) + 1
    parts = [f"{len(fixes)} auto-fixes"] + [f"{k}: {n}" for k, n in sorted(byk.items())]
    place = [f for f in flags if f["kind"] == "placement"]
    lay = [f for f in flags if f["kind"] == "layout"]
    fmt = [f for f in flags if f["kind"] == "format"]
    tail = f"{len(place)} placement, {len(lay)} layout, {len(fmt)} format flags"
    return "; ".join(parts) + " | " + tail


# ---- review gate: propose -> review -> commit --------------------------------
# Nothing is written to the production sheet by the standardiser. propose()
# computes the safe fixes and writes them to a SEPARATE proposed copy plus a
# human-readable change report. The fixes only reach production when a person
# calls commit() after reading the report. This stops a silent auto-overwrite.

def propose(src, proposed_path):
    """Compute fixes, write a proposed copy (original untouched). Returns (fixes, flags)."""
    fixes, flags, wb = check(src, fix_inplace=True)
    wb.save(proposed_path)
    return fixes, flags


def write_report(report_path, batches):
    """batches: list of (label, fixes, flags). Writes the reviewable change report."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ch = wb.active; ch.title = "Proposed changes"
    ch.append(["Batch", "Tab", "Row", "Field", "Old value", "New value", "Reason"])
    fl = wb.create_sheet("Flags (need a person)")
    fl.append(["Batch", "Tab", "Row", "Kind", "Field", "Value", "Why"])
    for label, fixes, flags in batches:
        for f in fixes:
            ch.append([label, f["tab"], f["row"], f["col"].strip(),
                       f["old"], f["new"], f["why"]])
        for f in flags:
            fl.append([label, f["tab"], f.get("row", ""), f["kind"],
                       f.get("col", "").strip(), f.get("value", ""), f["why"]])
    for ws in (ch, fl):
        for c in ws[1]:
            c.font = Font(bold=True)
        ws.freeze_panes = "A2"
    wb.save(report_path)
    return report_path


def commit(proposed_path, production_path):
    """Promote an approved proposed copy to production. Call ONLY after sign-off."""
    import shutil
    shutil.copyfile(proposed_path, production_path)
    return production_path


if __name__ == "__main__":
    import sys
    cmd = sys.argv[1]
    if cmd == "commit":
        commit(sys.argv[2], sys.argv[3]); print("committed:", sys.argv[3])
    else:                                   # propose (default)
        src = cmd if cmd.endswith((".xlsx", ".xlsm")) else sys.argv[2]
        proposed = src.replace(".xlsx", "_proposed.xlsx")
        fx, fl = propose(src, proposed)
        write_report(src.replace(".xlsx", "_standardise_report.xlsx"), [("job", fx, fl)])
        print("STANDARDISE (propose, not committed):", summarise(fx, fl))
        print("  wrote", proposed, "and the change report; commit only after sign-off")
