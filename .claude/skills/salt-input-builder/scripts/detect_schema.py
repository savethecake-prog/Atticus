"""Auto-detect each tab's column map so the build/audit never hard-code columns.

This is where silent off-by-one column errors used to creep in. The detector
reads header labels, matches them to known roles by synonym, locates the spec
block and the trailing identity/meta block, and emits a schema profile plus an
explicit list of low-confidence guesses for the human to confirm at the gate.

Output profile (JSON):
{
  "template_signature": "...",
  "tabs": {
    "<sheet>": {
      "header_row": 1, "first_data_row": 2,
      "example_rows": [2],
      "roles": {"title": 1, "whats_in_box": 2, "brand": 26, ... ,
                 "spec_start": 3, "spec_end": 18,
                 "confidence": 33, "concat": 34, "notes": 35, "source": 36},
      "labels": {"3": "Fan size", ...},
      "uncertain": ["source: not found - will be created after notes"]
    }
  }
}
"""
from __future__ import annotations
import sys, re
from openpyxl import load_workbook
from common import norm, header_signature, save_json

# role -> list of header synonyms (normalised substring match)
SYNONYMS = {
    "title": ["title", "product title", "name"],
    "whats_in_box": ["what's in the box", "whats in the box", "in the box", "box contents"],
    "brand": ["brand"],
    "sku": ["sku", "stock code", "stock keeping"],
    "model": ["model", "model number", "part number", "mpn", "p/n"],
    "ean": ["ean", "barcode", "gtin"],
    "tsin": ["tsin"],
    "warranty": ["warranty"],
    "category": ["category"],
    "video": ["video", "youtube", "media"],
    "confidence": ["confidence"],
    "concat": ["concat", "concatenate", "listing output", "final listing", "combined"],
    "notes": ["notes", "check notes", "check-notes", "comments", "qa notes"],
    "source": ["source", "root source", "source (root)", "reference"],
}
# roles that mark the start of the trailing identity/meta block (spec ends before these)
META_ROLES = ["brand", "sku", "model", "ean", "tsin", "warranty", "category",
              "video", "confidence", "concat", "notes", "source"]


def _match_role(label):
    n = norm(label)
    if not n:
        return None
    for role, syns in SYNONYMS.items():
        for s in syns:
            if n == s or s in n:
                return role
    return None


def detect_tab(ws):
    # header row = first row with >=3 non-empty cells
    header_row = 1
    for r in range(1, min(ws.max_row, 10) + 1):
        filled = sum(1 for c in range(1, ws.max_column + 1) if ws.cell(r, c).value not in (None, ""))
        if filled >= 3:
            header_row = r
            break
    roles, labels, uncertain = {}, {}, []
    for c in range(1, ws.max_column + 1):
        lab = ws.cell(header_row, c).value
        if lab not in (None, ""):
            labels[str(c)] = str(lab)
        role = _match_role(lab)
        if role and role not in roles:
            roles[role] = c
    # spec block: from the column after whats_in_box (or title) up to the first meta column
    spec_start = (roles.get("whats_in_box", roles.get("title", 1)) + 1)
    meta_cols = [roles[r] for r in META_ROLES if r in roles]
    spec_end = (min(meta_cols) - 1) if meta_cols else ws.max_column
    if spec_end < spec_start:
        uncertain.append("spec block could not be bounded - confirm spec_start/spec_end")
    roles["spec_start"], roles["spec_end"] = spec_start, spec_end
    # source column: if absent, it will be created immediately after notes
    if "source" not in roles:
        if "notes" in roles:
            roles["source"] = roles["notes"] + 1
            uncertain.append(f"source column absent; will be created at col {roles['source']} (after notes)")
        else:
            uncertain.append("notes column not found; cannot place source column - confirm layout")
    # example/template rows: any row whose title cell starts with 'example'
    first_data_row = header_row + 1
    example_rows = []
    tcol = roles.get("title", 1)
    for r in range(first_data_row, min(ws.max_row, first_data_row + 5) + 1):
        if norm(ws.cell(r, tcol).value).startswith("example"):
            example_rows.append(r)
    for need in ("title", "confidence", "notes"):
        if need not in roles:
            uncertain.append(f"role '{need}' not detected by header - confirm its column")
    return {
        "header_row": header_row,
        "first_data_row": first_data_row,
        "example_rows": example_rows,
        "roles": roles,
        "labels": labels,
        "uncertain": uncertain,
    }


def detect(path, out_path=None):
    wb = load_workbook(path)
    profile = {"template": path, "tabs": {}}
    sigs = []
    for ws in wb.worksheets:
        if ws.max_row < 1:
            continue
        tab = detect_tab(ws)
        tab["signature"] = header_signature(ws, tab["header_row"])
        sigs.append(tab["signature"])
        profile["tabs"][ws.title] = tab
    profile["template_signature"] = "|".join(sigs)
    if out_path:
        save_json(out_path, profile)
    return profile


if __name__ == "__main__":
    p = detect(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
    for name, t in p["tabs"].items():
        print(f"\n[{name}] header_row={t['header_row']} examples={t['example_rows']}")
        print("  roles:", {k: v for k, v in t["roles"].items()})
        if t["uncertain"]:
            print("  UNCERTAIN:", t["uncertain"])
