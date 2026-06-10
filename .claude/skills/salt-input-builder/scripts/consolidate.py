"""Tranche consolidation (W7): one tab per category, columns unioned, no row lost.

Tranches are sourced and built separately (and in parallel), then merged into the
consolidated workbook. The Endorfy "Keyboards" + "Keyboards (again)" failure was
two tranches of the SAME category landing on two tabs because their column sets
differed. The fix: group sheets by category identity, UNION the columns (nothing
dropped - CLAUDE.md S7), and stack the rows aligned by header so a column a tranche
lacks is left blank, never shifted.
"""
from __future__ import annotations
import re
from openpyxl import load_workbook, Workbook
from common import nkey


def category_key(sheet_title):
    """Normalised category identity for a tab title: 'Keyboards' and
    'Keyboards (again)' collapse to the same key."""
    t = re.sub(r"\(.*?\)", "", str(sheet_title or "")).strip()
    return nkey(t)


def consolidate(sources, out_path):
    """sources: list of workbook paths. Merge same-category sheets into one tab,
    unioning columns by normalised header and stacking rows. Returns out_path."""
    cats = {}  # key -> {"title", "headers"[], "hindex"{nkey:pos}, "rows"[{nkey:val}]}
    for path in sources:
        wb = load_workbook(path, data_only=True)
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(h).strip() if h is not None else "" for h in rows[0]]
            cat = cats.setdefault(category_key(ws.title),
                                  {"title": ws.title, "headers": [], "hindex": {}, "rows": []})
            for h in headers:
                hk = nkey(h)
                if hk and hk not in cat["hindex"]:
                    cat["hindex"][hk] = len(cat["headers"])
                    cat["headers"].append(h)
            for r in rows[1:]:
                if all(v in (None, "") for v in r):
                    continue
                cat["rows"].append({nkey(headers[i]): r[i] for i in range(min(len(headers), len(r)))})
    out = Workbook(); out.remove(out.active)
    for cat in cats.values():
        ws = out.create_sheet(title=str(cat["title"])[:31])
        ws.append(cat["headers"])
        for row in cat["rows"]:
            ws.append([row.get(nkey(h)) for h in cat["headers"]])
    out.save(out_path)
    return out_path


if __name__ == "__main__":
    import sys
    print("consolidated:", consolidate(sys.argv[1:-1], sys.argv[-1]))
