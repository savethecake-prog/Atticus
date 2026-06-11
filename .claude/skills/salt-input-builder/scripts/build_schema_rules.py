"""Build a dependency-free schema rules table from Christopher's structured
ecommerce data schema workbook (vendored under references/schema/).

This is a DEV-TIME tool. It reads the .xlsx (needs openpyxl) and emits
schema_rules.json, which is committed and read at runtime by schema_spec.py
WITHOUT openpyxl. The point (Christopher 2026-06-11): validate against a
field-keyed table — "calling a database to check is far superior to checking
against rules in a skill" — instead of loose header-substring heuristics.

Each field row carries: FIELD_ID, FIELD_NAME, FIELD_DESCRIPTION, FIELD_EXAMPLES,
FIELD_DATA-TYPE, FIELD_VALIDATION. We parse the free-text validation into
structured rules where it is safe to, and keep the raw text for the rest. We do
NOT invent rules the sheet does not state.

Usage:  python build_schema_rules.py [path/to/schema.xlsx] [out.json]
"""
from __future__ import annotations
import json, os, re, sys

# normalise a label (header or field name) to a loose match key
def nkey(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


# pull the quoted example values out of a FIELD_EXAMPLES cell. The sheet uses
# curly quotes “ ” (and occasionally a stray straight quote); split on those.
_QUOTED = re.compile(r"[“”\"]([^“”\"]+)[“”\"]")
def parse_examples(cell):
    if not cell:
        return []
    vals = [v.strip() for v in _QUOTED.findall(str(cell))]
    return [v for v in vals if v]


# parse the FIELD_VALIDATION free text into structured rules where safe.
_LEN = re.compile(r"([≤≥=])\s*(\d+)\s*characters?", re.I)  # ≤ ≥ = N characters
_GLYPH = re.compile(r"(\S+)\s+is required", re.I)
def parse_validation(cell, examples, data_type):
    raw = str(cell or "").strip()
    rules = {}
    if raw:
        rules["raw"] = raw
    low = raw.lower()
    # character-length constraints (a cell may carry several)
    for op, n in _LEN.findall(raw):
        n = int(n)
        if op == "≤":
            rules["max_len"] = n
        elif op == "≥":
            rules["min_len"] = n
        elif op == "=":
            rules["eq_len"] = n
    # compaction / no internal spaces
    if any(t in low for t in ("no spac", "flatten", "no internal space", "without space")):
        rules["compact"] = True
    # single-select enum: allowed set is the example values
    if "single-select" in low or "single select" in low:
        rules["single_select"] = True
        if examples:
            rules["allowed"] = examples
    # a required glyph the value must contain (e.g. "⎓ is required")
    m = _GLYPH.search(raw)
    if m and m.group(1) not in ("It", "Accuracy", "It’s"):
        rules["required_glyph"] = m.group(1)
    # an explicitly banned token (e.g. "DO NOT use MT/s")
    mb = re.search(r"do not use\s+([A-Za-z0-9/]+)", raw, re.I)
    if mb:
        rules["banned"] = [mb.group(1)]
    # apostrophe discipline: U+0027 straight, never U+2019 curly
    if "u+0027" in low and "u+2019" in low:
        rules["apostrophe"] = "straight"
    # capitalisation discipline
    if "sentence case" in low:
        rules["case"] = "sentence"
    # enum data-type with a fixed example set but not flagged single-select
    if (data_type or "").strip().lower() == "enum" and examples:
        rules.setdefault("allowed", examples)
    return rules


def build(xlsx_path):
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True)
    fields = {}
    order = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # find the header row: the row whose first cell is FIELD_ID
        header_row = None
        for r in range(1, min(ws.max_row, 5) + 1):
            if str(ws.cell(r, 1).value or "").strip().upper() == "FIELD_ID":
                header_row = r
                break
        start = (header_row + 1) if header_row else 1
        col = {str(ws.cell(header_row, c).value or "").strip().upper(): c
               for c in range(1, ws.max_column + 1)} if header_row else {
               "FIELD_ID": 1, "FIELD_NAME": 2, "FIELD_DESCRIPTION": 3,
               "FIELD_EXAMPLES": 4, "FIELD_DATA-TYPE": 5, "FIELD_VALIDATION": 6}
        def g(r, name):
            c = col.get(name)
            return ws.cell(r, c).value if c else None
        for r in range(start, ws.max_row + 1):
            fid = g(r, "FIELD_ID")
            if not fid or not str(fid).strip():
                continue
            fid = str(fid).strip()
            name = str(g(r, "FIELD_NAME") or "").strip()
            examples = parse_examples(g(r, "FIELD_EXAMPLES"))
            dtype = str(g(r, "FIELD_DATA-TYPE") or "").strip()
            rules = parse_validation(g(r, "FIELD_VALIDATION"), examples, dtype)
            rec = {
                "field_id": fid,
                "field_name": name,
                "name_key": nkey(name),
                "data_type": dtype,
                "examples": examples,
                "rules": rules,
                "source_tab": sheet,
            }
            if fid not in fields:           # first definition wins; dedupe repeats
                order.append(fid)
            fields[fid] = rec
    return [fields[fid] for fid in order]


def main():
    here = os.path.dirname(os.path.abspath(__file__))
    default_xlsx = os.path.join(here, "..", "references", "schema",
                                "ecommerce_schema_2026-06-11.xlsx")
    xlsx = sys.argv[1] if len(sys.argv) > 1 else default_xlsx
    out = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
        here, "..", "references", "schema", "schema_rules.json")
    table = build(xlsx)
    with open(out, "w", encoding="utf-8") as f:
        json.dump({"source": os.path.basename(xlsx), "fields": table},
                  f, ensure_ascii=False, indent=1)
    print(f"wrote {out}: {len(table)} fields")


if __name__ == "__main__":
    main()
