"""Completeness gate: catch products that surface too few specs.

The core's other guards stop INVENTION (no unsourced values, no invented fields).
This guards the opposite, equally costly failure: a product written with far fewer
specs than it should carry, which starves the downstream generator and produces a
low-value listing. It found nothing before because the skill had no guard against
under-capture at all.

It is agnostic. The baseline is learned from the data itself - the richest sibling
product in the same tab - and, when the operator records it, the number of fields
the source spec table actually published (schema tab key 'source_field_count').
Nothing here knows what any specific product "should" have.
"""
from __future__ import annotations
from openpyxl import load_workbook

# columns that are identity, prose, or role plumbing - not spec attributes
IDROLE = ("title", "subtitle", "brand", "sku", "ean", "tsin", "barcode", "categor",
          "warrant", "what", "confidence", "concat", "check", "source", "accuracy",
          "root", "image", "price", "model code", "product code", "model number", "added-from")


def _is_example(t):
    return str(t or "").strip().lower().startswith("example")


def _spec_headers(ws):
    return {c: str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)
            if ws.cell(1, c).value and not any(k in str(ws.cell(1, c).value).lower() for k in IDROLE)}


def check(built_path, schema, ratio=0.6, floor=4, product_tabs=None):
    """Findings for thin rows and under-discovered tabs.

    A product row flags if its populated spec count is below ratio * the tab's
    richest row, or below an absolute floor. A tab flags if it set up far fewer
    spec columns than the source table published (when that count was recorded).
    """
    wb = load_workbook(built_path, read_only=True, data_only=True)
    tabs = product_tabs or [t for t in schema["tabs"] if t in wb.sheetnames]
    findings = []
    for tab in tabs:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        spec_cols = _spec_headers(ws)
        rows = []
        for r in range(2, ws.max_row + 1):
            t = ws.cell(r, 1).value
            if t in (None, "") or _is_example(t):
                continue
            n = sum(1 for c in spec_cols if ws.cell(r, c).value not in (None, ""))
            rows.append((r, str(t), n))
        if not rows:
            continue
        richest = max(n for _, _, n in rows)
        src = schema["tabs"][tab].get("source_field_count") or 0
        # the expectation is the larger of the richest sibling and what the source
        # published; using the source count is what catches a whole batch being thin.
        baseline = max(richest, src)
        if baseline < floor:
            continue  # not a spec-bearing product tab (e.g. an instructions sheet)
        # lone-product blind spot: with a single product and no recorded source size,
        # the richest-sibling baseline is the row itself, so thinness can never fire.
        # Make that unmeasurable case loud instead of silently passing it.
        if not src and len(rows) <= 1:
            findings.append({"tab": tab, "row": None, "product": "(tab)",
                             "spec_count": rows[0][2], "severity": "SOFT",
                             "why": "completeness unverifiable: one product and no source_field_count "
                                    "recorded - capture the source table's field count so a thin lone "
                                    "product is caught (a single row cannot baseline against a sibling)"})
        for r, t, n in rows:
            why = []
            if n < ratio * baseline:
                ref = "the source table" if src and src >= richest else f"the richest {tab} row"
                why.append(f"{n} specs vs {baseline} in {ref}")
            if n < floor:
                why.append(f"only {n} specs (floor {floor})")
            if why:
                findings.append({"tab": tab, "row": r, "product": t[:50],
                                 "spec_count": n, "severity": "SOFT", "why": "; ".join(why)})
        if src and len(spec_cols) < ratio * src:
            findings.append({"tab": tab, "row": None, "product": "(tab)",
                             "spec_count": len(spec_cols), "severity": "SOFT",
                             "why": f"{len(spec_cols)} spec columns set up vs {src} fields in the "
                                    f"source table - discovery likely incomplete; capture the full table"})
    wb.close()
    return findings


def column_coverage(built_path, schema, mostly=0.5, default_structural=("tsin",)):
    """Coverage the row/tab checks miss, and the source of two real escapes
    (2026-06-10): (1) a whole target column never filled on any product - a
    never-attempted blank laundered as "unpublished"; (2) a row blank on a column
    its siblings fill - an identity/spec gap (e.g. the Xiaomi block with no SKU/EAN
    while every brand row has them). Structural-blank columns (a pre-listing TSIN)
    are skipped. The gate does not decide WHY a cell is empty; it refuses to let an
    empty pass silently, forcing: fill it, record a sourced-absent WITH evidence, or
    justify. No absence without a search.
    """
    import re
    nk = lambda s: re.sub(r"[^a-z0-9]", "", str(s or "").lower())
    wb = load_workbook(built_path, read_only=True, data_only=True)
    findings = []
    for tab in [t for t in schema.get("tabs", {}) if t in wb.sheetnames]:
        ws = wb[tab]
        t = schema["tabs"][tab]
        struct = set(default_structural) | {nk(s) for s in (t.get("gap_classes", {}) or {}).get("structural_blank", ())}
        titlec = t.get("roles", {}).get("title", 1)
        first = t.get("first_data_row", 2)
        examples = set(t.get("example_rows", []))
        headers = {c: str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
        prod = [r for r in range(first, ws.max_row + 1)
                if r not in examples and ws.cell(r, titlec).value not in (None, "")
                and not str(ws.cell(r, titlec).value).strip().lower().startswith("example")]
        n = len(prod)
        if not n:
            continue
        for c, h in headers.items():
            if nk(h) in struct:
                continue
            filled = sum(1 for r in prod if ws.cell(r, c).value not in (None, ""))
            if filled == 0:
                findings.append({"tab": tab, "col": h, "kind": "empty-column", "row": None,
                                 "why": f"column '{h}' is empty on all {n} products - source it, record a "
                                        f"sourced-absent WITH evidence, or justify; never leave a never-attempted "
                                        f"blank as if it were unpublished"})
            elif filled < n and filled / n >= mostly:
                holes = [r for r in prod if ws.cell(r, c).value in (None, "")]
                findings.append({"tab": tab, "col": h, "kind": "coverage-hole", "row": holes[0],
                                 "why": f"column '{h}' is filled for {filled}/{n} products but blank for "
                                        f"{len(holes)} (rows {holes}) - coverage gap vs siblings; source or justify"})
    wb.close()
    return findings


def coverage_closure(built_path, schema, entries, default_structural=("tsin",)):
    """Delivery gate - the symmetric twin of "no value without a source".

    Every blank target cell must be CLOSED: by a value, by a sourced 'absent' (a "No"
    justified by the complete table), or by a 'deferred' record carrying a search
    receipt (we tried, it is vendor-only / not public). A blank with no closing record
    is 'unknown' and UNRESOLVED - it may not ship, and it may not be declared
    absent/unpublished/vendor-only, without a recorded search. This closes the leak
    where a negative claim was made by reasoning instead of evidence. Structural blanks
    (a pre-listing TSIN) are exempt. Returns findings for unresolved blanks.
    """
    import re
    nk = lambda s: re.sub(r"[^a-z0-9]", "", str(s or "").lower())
    closed = {(e.get("tab"), e.get("row"), e.get("column"))
              for e in entries if e.get("answer_kind") in ("absent", "deferred")}
    wb = load_workbook(built_path, read_only=True, data_only=True)
    findings = []
    for tab in [t for t in schema.get("tabs", {}) if t in wb.sheetnames]:
        ws = wb[tab]
        t = schema["tabs"][tab]
        struct = set(default_structural) | {nk(s) for s in (t.get("gap_classes", {}) or {}).get("structural_blank", ())}
        titlec = t.get("roles", {}).get("title", 1)
        first = t.get("first_data_row", 2)
        examples = set(t.get("example_rows", []))
        headers = {c: str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
        for r in range(first, ws.max_row + 1):
            tv = ws.cell(r, titlec).value
            if r in examples or tv in (None, "") or str(tv).strip().lower().startswith("example"):
                continue
            for c, h in headers.items():
                if nk(h) in struct:
                    continue
                if ws.cell(r, c).value in (None, "") and (tab, r, c) not in closed:
                    findings.append({"tab": tab, "row": r, "col": h, "kind": "unresolved-blank",
                                     "why": f"'{h}' is blank with no closing record - fill it, or record a "
                                            f"sourced 'absent' (No) or a 'deferred' with a search receipt; "
                                            f"a blank cannot be declared absent/vendor-only without a search"})
    wb.close()
    return findings


def summarise(findings):
    rows = [f for f in findings if f["row"]]
    tabsf = [f for f in findings if f["row"] is None]
    tabs = sorted({f["tab"] for f in rows})
    parts = []
    if rows:
        parts.append(f"{len(rows)} thin product rows across {len(tabs)} tab(s)")
    if tabsf:
        parts.append(f"{len(tabsf)} tab(s) under-discovered vs source")
    return "; ".join(parts) if parts else "no thin rows flagged"


if __name__ == "__main__":
    import sys, os, common
    jobdir = sys.argv[1] if len(sys.argv) > 1 else "."
    schema = common.load_json(os.path.join(jobdir, "schema.json"))
    f = check(os.path.join(jobdir, "out.xlsx"), schema)
    print("COMPLETENESS:", summarise(f))
    for x in f:
        loc = f"{x['tab']} row{x['row']}" if x["row"] else x["tab"]
        print(f"  - {loc}: {x['why']}")
