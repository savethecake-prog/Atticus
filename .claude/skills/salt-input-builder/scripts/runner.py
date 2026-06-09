"""Runner: the thin driver that ties the phases together.

- status(jobdir): which artefacts exist, and the phase to resume from.
- needs_you(...): one consolidated review queue (validation, conflicts, dead
  links, gaps, sourceless rows) so the single adaptive pause has one list.
- delivery_summary(...): the hand-over statement (coverage by tier, gaps, flags).

CLI:
    python scripts/runner.py status  <jobdir>
    python scripts/runner.py needs    <jobdir>
    python scripts/runner.py summary  <jobdir>
A jobdir is expected to hold schema.json, ledger.json, and (after build) out.xlsx.
"""
from __future__ import annotations
import os, sys
from openpyxl import load_workbook
import common, ledger as L, confidence, completeness, standardise

IDENTITY = ["brand", "sku", "model", "ean", "tsin", "warranty", "category", "video"]


def _sourced_cols(roles):
    cols = set(range(roles["spec_start"], roles["spec_end"] + 1))
    if "whats_in_box" in roles:
        cols.add(roles["whats_in_box"])
    for r in IDENTITY:
        if r in roles:
            cols.add(roles[r])
    return cols


def status(jobdir):
    have = {f: os.path.exists(os.path.join(jobdir, f)) for f in
            ("job-spec.yaml", "schema.json", "ledger.json", "out.xlsx", "audit_report.md")}
    if not have["job-spec.yaml"]:
        nxt = "contract: generate and confirm the job spec"
    elif not have["schema.json"]:
        nxt = "schema profiling: detect and confirm the template column map"
    elif not have["ledger.json"]:
        nxt = "inspect-and-profile, then sourcing: build the category profile and the evidence ledger"
    elif not have["out.xlsx"]:
        nxt = "build: run build.py from the validated ledger"
    elif not have["audit_report.md"]:
        nxt = "audit: run audit.py with the confirmed category profile"
    else:
        nxt = "delivery: review the needs-you queue and the delivery summary"
    return {"have": have, "next": nxt}


def needs_you(schema, entries, built_path=None):
    q = []
    for v in L.validate(entries):
        q.append("ledger: " + v)
    for tab, row, col, field, sev, msg in L.detect_conflicts(entries):
        if sev == "HARD":
            q.append(f"conflict {tab} r{row}: {msg}")
    for e in entries:
        if e.get("source_url"):
            if e.get("link_ok") is False:
                q.append(f"link DEAD {e['tab']} r{e['row']} {e['field']}: {e['source_url']}")
            elif e.get("link_ok") is None and e.get("provenance") in ("manufacturer", "retailer"):
                q.append(f"link UNVERIFIED {e['tab']} r{e['row']} {e['field']}: {e['source_url']}")
    if built_path and os.path.exists(built_path):
        wb = load_workbook(built_path)
        for tab, t in schema.get("tabs", {}).items():
            if tab not in wb.sheetnames:
                continue
            ws = wb[tab]; roles = t["roles"]; cols = _sourced_cols(roles)
            cols |= set((t.get("extra_fields") or {}).values())
            for r in range(t["first_data_row"], ws.max_row + 1):
                if r in t.get("example_rows", []) or ws.cell(r, roles.get("title", 1)).value in (None, ""):
                    continue
                blank = [c for c in cols if ws.cell(r, c).value in (None, "")]
                if blank:
                    q.append(f"gap {tab} r{r}: {len(blank)} sourced cell(s) blank (left empty, not guessed)")
    return q


def delivery_summary(schema, entries, built_path=None):
    by_tier = {}
    for e in entries:
        by_tier[e["provenance"]] = by_tier.get(e["provenance"], 0) + 1
    conf = L.detect_conflicts(entries)
    hard = sum(1 for c in conf if c[4] == "HARD")
    soft = len(conf) - hard
    dead = sum(1 for e in entries if e.get("link_ok") is False)
    unver = sum(1 for e in entries if e.get("link_ok") is None and e.get("provenance") in ("manufacturer", "retailer"))
    nq = needs_you(schema, entries, built_path)
    gaps = sum(1 for n in nq if n.startswith("gap "))
    cmap = confidence.export_map(entries)
    held = [c for c in cmap if not c["eligible"]]
    avg = round(sum(c["confidence"] for c in cmap) / len(cmap)) if cmap else 0
    thin = "not checked (build first)"
    if built_path:
        try:
            thin = completeness.summarise(completeness.check(built_path, schema))
        except Exception as e:
            thin = f"check failed: {e}"
    L_ = ["# Delivery summary", "",
          f"- ledger entries: {len(entries)}",
          "- by provenance tier: " + ", ".join(f"{k}={v}" for k, v in sorted(by_tier.items())),
          f"- values written: {len(cmap)} (mean confidence {avg}/100); flagged for accuracy check: {len(held)}",
          f"- source conflicts: {hard} hard, {soft} soft",
          f"- links: {dead} dead, {unver} unverified",
          f"- gaps (blank sourced cells, left empty not guessed): {gaps}",
          f"- completeness: {thin}",
          f"- open needs-you items: {len(nq)}",
          "",
          "This data is fully sourced, traceable to a snippet, independently audited,",
          "and its gaps are marked. It is not asserted to be correct - it is asserted",
          "to be honest about where every value came from and where it is missing."]
    return "\n".join(L_) + "\n"


if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else "status"
    jobdir = sys.argv[2] if len(sys.argv) > 2 else "."
    if cmd == "status":
        s = status(jobdir)
        print("artefacts:", {k: ("yes" if v else "no") for k, v in s["have"].items()})
        print("next:", s["next"])
    elif cmd == "standardise":
        # review gate: propose fixes + write the change report; never auto-commit
        built = os.path.join(jobdir, "out.xlsx")
        proposed = os.path.join(jobdir, "out_proposed.xlsx")
        report = os.path.join(jobdir, "standardise_report.xlsx")
        fx, fl = standardise.propose(built, proposed)
        standardise.write_report(report, [("job", fx, fl)])
        print("STANDARDISE (propose, NOT committed):", standardise.summarise(fx, fl))
        print(f"  review {os.path.basename(report)} and {os.path.basename(proposed)};")
        print("  then: python scripts/standardise.py commit out_proposed.xlsx out.xlsx")
    else:
        schema = common.load_json(os.path.join(jobdir, "schema.json"))
        entries = L.load(os.path.join(jobdir, "ledger.json"))
        built = os.path.join(jobdir, "out.xlsx")
        if cmd == "needs":
            q = needs_you(schema, entries, built)
            print("NEEDS YOU (%d):" % len(q))
            for n in q:
                print("  -", n)
        elif cmd == "summary":
            print(delivery_summary(schema, entries, built))
        elif cmd == "completeness":
            f = completeness.check(built, schema)
            print("COMPLETENESS:", completeness.summarise(f))
            for x in f:
                loc = f"{x['tab']} row{x['row']}" if x["row"] else x["tab"]
                print(f"  - {loc}: {x['why']}")
        elif cmd == "confidence":
            import json
            cmap = confidence.export_map(entries)
            json.dump(cmap, open(os.path.join(jobdir, "confidence_map.json"), "w"), indent=2)
            print(f"wrote confidence_map.json ({len(cmap)} cells) for SALT to gate on")
