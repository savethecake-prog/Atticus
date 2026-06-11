"""Runtime schema validator — the field-keyed "tester" Christopher asked for.

Reads schema_rules.json (built by build_schema_rules.py from the vendored
workbook) WITHOUT openpyxl, maps each house header to its schema field by a
normalised name key — not a loose substring — and checks the value against the
field's parsed rules. It VALIDATES and REPORTS; it never rewrites a value (that
is the standardiser's job). This is the harness a tester agent calls; the model
does not self-assess (Christopher 2026-06-11: models are poor at verification).

A value is only validated where we can confidently key it to a field. No mapping
-> no rule -> no finding. We never invent a constraint the schema does not state.
"""
from __future__ import annotations
import json, os, re

_RULES_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           "..", "references", "schema", "schema_rules.json")

_CURLY_APOS = "’"   # ’  (U+2019) — banned in What's-in-the-box
_STRAIGHT = "'"     # '  (U+0027)


def nkey(s) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


class Schema:
    def __init__(self, fields):
        self.fields = fields
        self.by_id = {f["field_id"]: f for f in fields}
        # name_key -> field (first definition wins, matching the build order)
        self.by_name = {}
        for f in fields:
            self.by_name.setdefault(f["name_key"], f)

    def lookup(self, header):
        """Resolve a house header (e.g. 'Video memory capacity:') to its field.
        Keys on the normalised name, so 'video' no longer collides with media."""
        k = nkey(header)
        if not k:
            return None
        f = self.by_name.get(k)
        if f:
            return f
        # tolerate a trailing-word mismatch: exact key only, else None. We do not
        # guess; an unmapped header simply carries no schema rules.
        return None

    def validate_value(self, header, value):
        """Return a list of violation dicts for one cell. Empty == clean/unmapped."""
        f = self.lookup(header)
        if f is None or value is None:
            return []
        v = value if isinstance(value, str) else str(value)
        if v.strip() == "":
            return []
        rules = f.get("rules", {})
        out = []
        def viol(rule, why):
            out.append({"field_id": f["field_id"], "field_name": f["field_name"],
                        "rule": rule, "why": why, "value": v})
        n = len(v)
        if "max_len" in rules and n > rules["max_len"]:
            viol("max_len", f"{n} chars > max {rules['max_len']}")
        if "min_len" in rules and n < rules["min_len"]:
            viol("min_len", f"{n} chars < min {rules['min_len']}")
        if "eq_len" in rules and n != rules["eq_len"]:
            viol("eq_len", f"{n} chars != required {rules['eq_len']}")
        if rules.get("single_select") and rules.get("allowed"):
            allowed = {a.lower() for a in rules["allowed"]}
            if v.strip().lower() not in allowed:
                viol("single_select", f"{v!r} not in allowed {rules['allowed']}")
        if rules.get("compact") and re.search(r"\s", v):
            viol("compact", "value must have no internal spaces")
        if "required_glyph" in rules and rules["required_glyph"] not in v:
            viol("required_glyph", f"missing required {rules['required_glyph']!r}")
        for bad in rules.get("banned", []):
            if bad.lower() in v.lower():
                viol("banned", f"contains banned token {bad!r}")
        if rules.get("apostrophe") == "straight" and _CURLY_APOS in v:
            viol("apostrophe", f"uses curly U+2019 ’; must be straight U+0027 '")
        if rules.get("case") == "sentence" and _is_shouty(v):
            viol("case", "should be sentence case, not title/all-caps")
        return out


def _is_shouty(v):
    """Conservative all-caps / title-case detector for prose fields: flags only a
    clearly non-sentence-case string, never a normal sentence."""
    letters = [c for c in v if c.isalpha()]
    if len(letters) < 4:
        return False
    if all(c.isupper() for c in letters):
        return True
    return False


def load(path=None):
    with open(path or _RULES_PATH, encoding="utf-8") as fh:
        data = json.load(fh)
    return Schema(data["fields"])


def validate_workbook(xlsx_path, schema=None):
    """Validate every mapped cell in a built workbook. Returns a list of findings.
    Requires openpyxl only here (workbook IO), not for value-level validation."""
    from openpyxl import load_workbook
    schema = schema or load()
    wb = load_workbook(xlsx_path)
    findings = []
    for ws in wb.worksheets:
        headers = {c: str(ws.cell(1, c).value or "")
                   for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value in (None, ""):
                continue
            for c, hdr in headers.items():
                for f in schema.validate_value(hdr, ws.cell(r, c).value):
                    findings.append({"tab": ws.title, "row": r, **f})
    return findings


if __name__ == "__main__":
    import sys
    sch = load()
    if len(sys.argv) > 1:
        fnd = validate_workbook(sys.argv[1], sch)
        print(f"{len(fnd)} schema violations")
        for x in fnd[:50]:
            print(f"  {x['tab']} r{x['row']} {x['field_name']}: {x['rule']} — {x['why']}")
    else:
        print(f"schema_rules: {len(sch.fields)} fields loaded")
