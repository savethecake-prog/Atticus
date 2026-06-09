"""Shared helpers for the salt-input-builder core.

Nothing here knows anything about a specific client, brand, or template.
It only knows how to read/write cells, colour by provenance tier, sign a
header layout, and normalise text for value-vs-snippet matching.
"""
from __future__ import annotations
import json, re, hashlib, copy
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---- provenance tiers -------------------------------------------------------
# A value's provenance is one of these. "client" = supplied by the user's own
# inventory/PIM (their SKU, their category); "derived" = a formula or computed
# cell. A value is NEVER allowed to be "model" / from memory.
TIERS = ("manufacturer", "retailer", "client", "derived")
WEB_TIERS = ("manufacturer", "retailer")  # these require url + snippet

# Default confidence colours (fill, font). Override via job spec.
DEFAULT_COLOURS = {
    "manufacturer": {"fill": "C6EFCE", "font": "006100"},  # green
    "retailer":     {"fill": "FFEB9C", "font": "9C6500"},  # orange
    "client":       {"fill": "DDEBF7", "font": "1F4E78"},  # light blue (informational)
    "derived":      {"fill": "F2F2F2", "font": "595959"},  # grey
    "unverifiable": {"fill": "FFC7CE", "font": "9C0006"},  # red (blank+flag / discrepancy)
}

BODY_FONT = dict(name="Arial", size=10)
TOP_WRAP = dict(vertical="top", wrap_text=True)


def fill(hex6):
    return PatternFill("solid", fgColor="FF" + hex6.upper())


def tier_rank(tier):
    """Lower rank = stronger. Used to pick a row's overall (most conservative) tier."""
    return {"manufacturer": 0, "retailer": 1, "client": 2, "derived": 3}.get(tier, 9)


def write_cell(ws, row, col, value, *, font_hex=None):
    c = ws.cell(row, col, value=value)
    c.font = Font(color=("FF" + font_hex if font_hex else None), **BODY_FONT)
    c.alignment = Alignment(**TOP_WRAP)
    return c


def set_fill(ws, row, col, hex6):
    ws.cell(row, col).fill = fill(hex6)


def col_letter(i):
    return get_column_letter(i)


# ---- text normalisation for value-vs-snippet matching -----------------------
def norm(s):
    """Lowercase, collapse whitespace, strip common unit punctuation noise."""
    s = "" if s is None else str(s)
    s = s.replace("\u00a0", " ").replace("×", "x").replace("≤", "<=").replace("≥", ">=")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def value_supported_by_snippet(value, snippet):
    """True if the written value is traceable to the snippet.

    Conservative: every numeric token in the value must appear in the snippet,
    and at least half the alphanumeric word tokens must appear. This catches
    transcription drift (a wrong number) without demanding a verbatim copy.
    """
    v, s = norm(value), norm(snippet)
    if not v:
        return True
    nums_v = re.findall(r"\d+\.?\d*", v)
    nums_s = set(re.findall(r"\d+\.?\d*", s))
    for n in nums_v:
        if n not in nums_s:
            return False
    words_v = [w for w in re.findall(r"[a-z0-9]+", v) if not w.isdigit()]
    if not words_v:
        return True
    hits = sum(1 for w in words_v if w in s)
    return hits >= max(1, len(words_v) // 2)


# ---- header-layout signature ------------------------------------------------
def header_signature(ws, header_row=1):
    """SHA1 of the (column -> header text) layout for a sheet. If a template's
    signature changes, the saved schema profile must be re-confirmed."""
    cells = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        cells.append(f"{c}:{norm(v)}")
    return hashlib.sha1("|".join(cells).encode()).hexdigest()[:12]


# ---- IO ---------------------------------------------------------------------
def load_json(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def save_json(path, obj):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, indent=2, ensure_ascii=False)
