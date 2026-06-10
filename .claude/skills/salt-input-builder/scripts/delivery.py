"""Produce the delivery copy (W2).

The working/analysis workbook carries Source, Confidence and Check-notes columns,
confidence shading and a key sheet - all of which help us, none of which belong on
the sheet handed to SALT. The delivery copy strips those, drops the hidden
(all-blank) columns so no dead column ships, leaves every unknown cell truly blank
(there is no marker text in the new model), and applies the house Excel format. It
never changes a value.
"""
from __future__ import annotations
from openpyxl import load_workbook
import format_xlsx

# header substrings that are analysis plumbing, not listing data
_ANALYSIS = ("source (root)", "source root", "confidence", "check notes",
             "check note", "accuracy")
_ANALYSIS_SHEETS = ("confidence key",)


def _is_analysis(header):
    h = str(header or "").strip().lower()
    return any(k in h for k in _ANALYSIS)


def make_delivery(src, out_path, schema=None, entries=None, drop_hidden=True):
    """Write a clean delivery copy of src: remove analysis columns and sheets, drop
    hidden (all-blank) columns, apply the house format. Returns out_path.

    Enforcement: if a schema is supplied, delivery REFUSES (raises) when
    completeness.coverage_closure finds an unresolved blank - a cell that is neither
    filled, sourced-absent, nor a receipted deferred. A blank may not ship, nor be
    declared absent/vendor-only, without a recorded search. (build.py refuses on a
    dirty ledger; this is the same wall at the delivery boundary.)"""
    if schema is not None:
        import completeness
        unresolved = completeness.coverage_closure(src, schema, entries or [])
        if unresolved:
            u = unresolved[0]
            raise ValueError(
                f"delivery refused: {len(unresolved)} unresolved blank(s) - fill, record a sourced "
                f"'absent' (No), or receipt as a 'deferred' with a search receipt before delivery. "
                f"First: {u['tab']} r{u['row']} '{u['col']}'")
    wb = load_workbook(src)
    for name in list(wb.sheetnames):
        if name.strip().lower() in _ANALYSIS_SHEETS:
            del wb[name]
    for ws in wb.worksheets:
        # right-to-left so deletions do not shift the columns still to be checked
        for c in range(ws.max_column, 0, -1):
            hdr = ws.cell(1, c).value
            hidden = ws.column_dimensions[ws.cell(1, c).column_letter].hidden
            if _is_analysis(hdr) or (drop_hidden and hidden):
                ws.delete_cols(c, 1)
    wb.save(out_path)
    format_xlsx.format_workbook(out_path)
    return out_path


if __name__ == "__main__":
    import sys
    print("delivery:", make_delivery(sys.argv[1], sys.argv[2]))
