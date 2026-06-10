"""The closure reflex - the deterministic parts.

When the coverage gate (completeness.coverage_closure) finds unresolved blanks, the
orchestrator runs a loop: gate -> brief -> fan out one sourcer per gap-product ->
merge -> re-gate -> receipt the residue. So a negative is never asserted by reasoning;
the search runs first, and what it can't fill is closed with a receipt of the attempt.

These helpers are the deterministic spine; the per-product sourcing fan-out is the
orchestrator's (an agent per product, manufacturer/source-tier first). HUMAN-IN-THE-LOOP
is preserved by construction: identity fields (barcode/SKU/TSIN/model number) are NEVER
auto-sourced from the web - they are DEFERRED to the distributor master with a receipt
and a human confirm (the GEX750 discipline). Only spec fields are auto-sourced.
"""
from __future__ import annotations
import re
from openpyxl import load_workbook
from common import classify_gap
import standardise

# value must come from the distributor master, confirmed by a human - never web-asserted
IDENTITY_FIELDS = ("sku", "ean", "barcode", "tsin", "model number", "model code", "model no")
_IDENT = {re.sub(r"[^a-z0-9]", "", f) for f in IDENTITY_FIELDS}


def _nk(s):
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def sourcing_brief(built_path, schema, closure_findings):
    """Group the gate's unresolved cells into a per-product task list, DERIVED from the
    gate (which is derived from the schema) so it can never silently omit a column.
    Each gap field is split: 'auto_source' (web-targetable spec) or 'defer' (identity /
    region-variable -> master, never web-asserted). Returns {(tab,row): {identity,
    auto_source[], defer[(field,reason)]}}."""
    wb = load_workbook(built_path, read_only=True, data_only=True)
    briefs = {}
    ident_keys = {_nk(x) for x in ("brand", "model", "model number", "product title")}
    for f in closure_findings:
        tab, row, field = f["tab"], f["row"], f["col"]
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        gc = (schema["tabs"].get(tab, {}).get("gap_classes", {}) or {})
        key = (tab, row)
        if key not in briefs:
            hdr = {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}
            ident = {h: ws.cell(row, c).value for h, c in hdr.items()
                     if _nk(h) in ident_keys and ws.cell(row, c).value not in (None, "")}
            briefs[key] = {"identity": ident, "auto_source": [], "defer": []}
        if _nk(field) in _IDENT:
            briefs[key]["defer"].append((field, "identity - from the distributor master, not the web (GEX750); human-confirm"))
            continue
        kind = classify_gap(field, derivable_fields=gc.get("derivable", ()),
                            region_variable_fields=gc.get("region_variable", ()),
                            structural_blank_fields=gc.get("structural_blank", ()))
        if kind == "region_variable":
            briefs[key]["defer"].append((field, "region-variable - confirm from master/region"))
        else:
            briefs[key]["auto_source"].append(field)
    wb.close()
    return briefs


def merge_sourced(built_path, results, schema, out_path=None):
    """Fill BLANK target cells from sourced results, keyed by the row's MODEL NUMBER
    (variant-safe - a model number is unique, unlike a name that strips '+'). Standardises
    each value, never overrides a filled cell. results: {model_number: {field: value | {value}}}.
    Returns the count filled."""
    wb = load_workbook(built_path)
    filled = 0
    for tab, t in schema.get("tabs", {}).items():
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        hdr = {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}
        mncol = next((c for h, c in hdr.items() if _nk(h) in ("modelnumber", "modelcode", "modelno")), None)
        if not mncol:
            continue
        for r in range(t.get("first_data_row", 2), ws.max_row + 1):
            mn = ws.cell(r, mncol).value
            rec = results.get(str(mn).strip()) if mn not in (None, "") else None
            if not rec:
                continue
            for field, obj in rec.items():
                c = hdr.get(field) or next((cc for h, cc in hdr.items() if _nk(h) == _nk(field)), None)
                if c and ws.cell(r, c).value in (None, ""):
                    val = obj.get("value") if isinstance(obj, dict) else obj
                    if val:
                        nv, _ = standardise.standardise_value(field, str(val))
                        ws.cell(r, c).value = nv
                        filled += 1
    wb.save(out_path or built_path)
    return filled


def receipt_residue(closure_findings, receipts):
    """Close each still-unresolved blank with a 'deferred' ledger entry carrying its
    search receipt (the searches that came back empty, or the defer reason). A finding
    with no receipt is returned in `unreceipted` - it CANNOT ship, the orchestrator must
    supply one. receipts: {(tab,row,column): text} or {(tab,row,normalised-field): text}.
    Returns (deferred_entries, unreceipted_findings)."""
    entries, unreceipted = [], []
    for f in closure_findings:
        rcpt = (receipts.get((f["tab"], f["row"], f.get("column")))
                or receipts.get((f["tab"], f["row"], _nk(f.get("col")))))
        if not rcpt:
            unreceipted.append(f)
            continue
        entries.append({"tab": f["tab"], "row": f["row"], "field": f.get("col"),
                        "column": f.get("column"), "answer_kind": "deferred", "search_receipt": rcpt})
    return entries, unreceipted
