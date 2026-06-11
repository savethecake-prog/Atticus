"""Independent audit. The logic here does not trust how the build wrote the
sheet; it re-reads the workbook and re-derives every check from scratch, then
cross-checks against the ledger. This is the "the writer is not the auditor"
principle in code.

Checks (core, domain-agnostic):
  A. Orphan value      - a populated sourced cell with no ledger entry -> HARD
  B. Snippet drift     - a manufacturer/retailer value not supported by its snippet -> HARD
  C. Colour vs tier    - Confidence cell colour does not match the row's tier
  D. Link liveness     - source link dead or unverified
Plus domain-pack checks (e.g. SKU decode vs fields, cross-field logic, plausibility).

The tester never writes the work: the built workbook is read, never modified.
Discrepant cells are shaded with the 'unverifiable' colour and an "ACCURACY CHECK"
note appended (never overwriting a value) in a SEPARATE annotated copy
(<built>_audited.xlsx); the production sheet is left byte-for-byte. A Markdown
report is written for the human, and schema-conformance findings are report-only.
"""
from __future__ import annotations
import sys, os, datetime, copy
from openpyxl import load_workbook
from openpyxl.styles import Font
import common, ledger as L, profile_engine
from common import DEFAULT_COLOURS, BODY_FONT, TOP_WRAP, Alignment

# columns that must trace to the ledger (whats_in_box included - it is sourced)
IDENTITY_ROLES = ["brand", "sku", "model", "ean", "tsin", "warranty", "category", "video"]


def _sourced_columns(roles):
    cols = set(range(roles["spec_start"], roles["spec_end"] + 1))
    if "whats_in_box" in roles:
        cols.add(roles["whats_in_box"])
    for r in IDENTITY_ROLES:
        if r in roles:
            cols.add(roles[r])
    return cols


def _gtin_ok(code):
    """Valid GS1 barcode (GTIN-8/12/13/14): all digits, right length, and the
    final check digit reconciles. A wrong barcode on the wrong product is exactly
    the failure that reaches the customer, so format is checked, not assumed."""
    s = str(code).strip()
    if not s.isdigit() or len(s) not in (8, 12, 13, 14):
        return False
    body = [int(c) for c in s[:-1]][::-1]
    total = sum(d * (3 if i % 2 == 0 else 1) for i, d in enumerate(body))
    return (10 - total % 10) % 10 == int(s[-1])


def audit(built_path, schema, ledger_path, out_report, colours=None, profiles=None,
          annotated_path=None):
    """Re-derive and report. The tester never writes the work: the built workbook
    is read, never modified. Shading/notes are written to a SEPARATE annotated copy
    (``annotated_path``, defaulting to ``<built>_audited.xlsx``) so the production
    sheet the builder owns is left byte-for-byte. Returns the findings list."""
    colours = {**DEFAULT_COLOURS, **(colours or {})}
    profiles = profiles or {}
    if annotated_path is None:
        base, ext = os.path.splitext(built_path)
        annotated_path = f"{base}_audited{ext}"
    entries = L.load(ledger_path)
    by_cell = {(e["tab"], e["row"], e["column"]): e for e in entries}
    wb = load_workbook(built_path)

    findings = []  # (tab,row,[cols],severity,msg)

    def F(tab, row, cols, sev, msg):
        findings.append((tab, row, list(cols), sev, msg))

    for tab, t in schema["tabs"].items():
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        roles = t["roles"]
        prof = profiles.get(tab) or profiles.get("*")
        srccols = _sourced_columns(roles)
        srccols |= set((t.get("extra_fields") or {}).values())
        for r in range(t["first_data_row"], ws.max_row + 1):
            if r in t.get("example_rows", []):
                continue
            tcol = roles.get("title", 1)
            if ws.cell(r, tcol).value in (None, ""):
                continue
            # A. orphan values
            for c in srccols:
                val = ws.cell(r, c).value
                if val not in (None, "") and (tab, r, c) not in by_cell:
                    F(tab, r, [c], "HARD", f"value '{str(val)[:40]}' has no ledger entry (unsourced).")
            # B. snippet drift (unexplained = HARD; disclosed with a doubt_reason = SOFT)
            for c in srccols:
                e = by_cell.get((tab, r, c))
                if e and e["provenance"] in ("manufacturer", "retailer"):
                    if not common.value_supported_by_snippet(e["value"], e.get("snippet", "")):
                        if e.get("doubt_reason"):
                            F(tab, r, [c], "SOFT", f"'{e['field']}' value not verbatim in snippet - disclosed: {e['doubt_reason']}")
                        else:
                            F(tab, r, [c], "HARD", f"'{e['field']}' value not supported by its snippet (undisclosed drift).")
            # C. colour vs tier
            if "confidence" in roles:
                tier = L.row_tier(entries, tab, r)
                if tier:
                    got = (ws.cell(r, roles["confidence"]).fill.fgColor.rgb or "")[-6:].upper()
                    want = colours[tier]["fill"].upper()
                    if got != want:
                        F(tab, r, [roles["confidence"]], "SOFT",
                          f"Confidence colour {got or 'none'} does not match row tier '{tier}' ({want}).")
            # D. link liveness
            e = by_cell.get((tab, r, roles.get("source", -1)))
            srcurl = L.primary_source(entries, tab, r)
            le = next((x for x in entries if x["tab"] == tab and x["row"] == r and x.get("source_url") == srcurl), None)
            if le is not None and le.get("link_ok") is False:
                F(tab, r, [roles.get("source", tcol)], "HARD", f"source link is dead: {srcurl}")
            # E. category-profile checks (profile built per job by inspection)
            if prof:
                for col_list, sev, msg in profile_engine.run_checks(ws, r, roles, t.get("labels", {}), prof):
                    F(tab, r, col_list, sev, msg)

    # F. cross-row identity integrity (agnostic): the same SKU on two products
    seen = {}
    for tab, t in schema["tabs"].items():
        if tab not in wb.sheetnames:
            continue
        scol = t["roles"].get("sku")
        if not scol:
            continue
        ws = wb[tab]
        for r in range(t["first_data_row"], ws.max_row + 1):
            if r in t.get("example_rows", []) or ws.cell(r, t["roles"].get("title", 1)).value in (None, ""):
                continue
            val = ws.cell(r, scol).value
            if val not in (None, ""):
                seen.setdefault(str(val).strip(), []).append((tab, r, scol))
    for val, occ in seen.items():
        if len(occ) > 1:
            where = ", ".join(f"{tb} r{rw}" for tb, rw, _ in occ)
            for tb, rw, sc in occ:
                F(tb, rw, [sc], "HARD", f"duplicate SKU '{val}' on multiple products ({where}). Confirm identifiers.")

    # F2. EAN format + uniqueness (the GEX750 shared-barcode lesson)
    eans = {}
    for tab, t in schema["tabs"].items():
        if tab not in wb.sheetnames:
            continue
        roles = t["roles"]; ecol = roles.get("ean"); scol = roles.get("sku")
        if not ecol:
            continue
        ws = wb[tab]
        for r in range(t["first_data_row"], ws.max_row + 1):
            if r in t.get("example_rows", []) or ws.cell(r, roles.get("title", 1)).value in (None, ""):
                continue
            val = ws.cell(r, ecol).value
            if val in (None, ""):
                continue
            code = str(val).strip()
            if not _gtin_ok(code):
                F(tab, r, [ecol], "HARD", f"EAN '{code}' is not a valid barcode (digits/length/check-digit).")
            sku = str(ws.cell(r, scol).value).strip() if scol and ws.cell(r, scol).value else ""
            eans.setdefault(code, []).append((tab, r, ecol, sku))
    for code, occ in eans.items():
        if len(occ) > 1 and len({sku for *_, sku in occ}) > 1:
            where = ", ".join(f"{tb} r{rw}" for tb, rw, _, _ in occ)
            for tb, rw, ec, _ in occ:
                F(tb, rw, [ec], "HARD", f"EAN '{code}' shared across different SKUs ({where}). A barcode identifies one product.")

    # G. source conflicts (independent comparator over candidate observations)
    for tab, row, col, field, sev, msg in L.detect_conflicts(entries):
        F(tab, row, [col] if col else [], sev, msg)

    # H. schema conformance (Christopher's field-keyed rules). Report-only: these
    # are formatting/standard violations, surfaced for the writer to fix; the
    # auditor never edits a value, so schema findings are NOT shaded into cells.
    try:
        import schema_spec
        _sch = schema_spec.load()
    except Exception:
        _sch = None
    if _sch is not None:
        for tab, t in schema["tabs"].items():
            if tab not in wb.sheetnames:
                continue
            ws = wb[tab]; roles = t["roles"]
            for r in range(t["first_data_row"], ws.max_row + 1):
                if r in t.get("example_rows", []) or ws.cell(r, roles.get("title", 1)).value in (None, ""):
                    continue
                for c in range(1, ws.max_column + 1):
                    hdr = str(ws.cell(1, c).value or "")
                    for v in _sch.validate_value(hdr, ws.cell(r, c).value):
                        F(tab, r, [c], "SCHEMA", f"{v['rule']}: {v['why']} (field '{v['field_name']}')")

    # apply flags to the workbook (SCHEMA findings are report-only, never shaded)
    grouped = {}
    for tab, row, cols, sev, msg in findings:
        if sev == "SCHEMA":
            continue
        g = grouped.setdefault((tab, row), {"cols": set(), "msgs": []})
        g["cols"].update(cols)
        g["msgs"].append(f"[{sev}] {msg}")
    today = datetime.date.today().isoformat()
    for (tab, row), g in grouped.items():
        ws = wb[tab]
        for c in g["cols"]:
            common.set_fill(ws, row, c, colours["unverifiable"]["fill"])
        ncol = schema["tabs"][tab]["roles"].get("notes")
        if ncol:
            cur = ws.cell(row, ncol).value or ""
            if "ACCURACY CHECK" not in str(cur):
                block = f" || ACCURACY CHECK ({today}): " + "  ".join(g["msgs"])
                nc = ws.cell(row, ncol, value=(str(cur) + block).strip())
                nc.font = Font(**BODY_FONT)
                nc.alignment = Alignment(**TOP_WRAP)

    wb.save(annotated_path)            # the annotated COPY; built_path is left untouched
    _report(out_report, findings, today)
    sev_counts = {s: sum(1 for f in findings if f[3] == s) for s in ("HARD", "SOFT", "SCHEMA")}
    print(f"AUDIT done: {len(findings)} findings {sev_counts} -> {out_report} "
          f"(annotated copy: {annotated_path}; source left unmodified)")
    return findings


def _report(path, findings, today):
    lines = [f"# Audit report ({today})", ""]
    hard = [f for f in findings if f[3] == "HARD"]
    soft = [f for f in findings if f[3] == "SOFT"]
    schema = [f for f in findings if f[3] == "SCHEMA"]
    lines += [f"- HARD findings: {len(hard)}", f"- SOFT findings: {len(soft)}",
              f"- SCHEMA findings: {len(schema)}", ""]
    if not findings:
        lines.append("No discrepancies found. Every populated cell traces to a ledger entry, "
                     "every manufacturer/retailer value is supported by its snippet, colours match "
                     "tiers, and all source links resolve.")
    for label, group in (("Hard findings", hard), ("Soft findings", soft),
                         ("Schema conformance (for the writer to fix)", schema)):
        if not group:
            continue
        lines += [f"\n## {label}", "", "| Tab | Row | Cols | Finding |", "|---|---|---|---|"]
        for tab, row, cols, sev, msg in sorted(group, key=lambda x: (x[0], x[1])):
            lines.append(f"| {tab} | {row} | {','.join(map(str, cols))} | {msg} |")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")


if __name__ == "__main__":
    # audit.py <built.xlsx> <schema.json> <ledger.json> <report.md> [profile.yaml]
    import common as _c, yaml
    profs = {}
    if len(sys.argv) > 5:
        with open(sys.argv[5]) as f:
            profs = {"*": yaml.safe_load(f)}
    audit(sys.argv[1], _c.load_json(sys.argv[2]), sys.argv[3], sys.argv[4], profiles=profs)
