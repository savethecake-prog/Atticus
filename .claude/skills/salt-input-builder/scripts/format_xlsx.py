"""House Excel formatting for the delivery copy (W6).

A standardised, legible layout so every delivered workbook reads the same way,
regardless of which client template or category it came from:

  - row height 30 on every row;
  - columns A (Title) and B (What's in the Box) width 45, all others 16;
  - every cell wrapped and vertically centred;
  - the header row centred and bold; body cells left-aligned;
  - thin borders across the whole data range.

It touches presentation only - never a value. Idempotent: re-running it produces
the same layout. This runs on the delivery copy, after standardise and after any
analysis columns have been stripped.
"""
from __future__ import annotations
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font

ROW_HEIGHT = 30
WIDE_COLS = {1: 45, 2: 45}          # A = Title, B = What's in the Box
DEFAULT_WIDTH = 16
_THIN = Side(style="thin", color="FFB0B0B0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def format_workbook(path, out_path=None):
    """Apply the house format to every sheet. In place unless out_path is given.
    Returns the path written."""
    wb = load_workbook(path)
    for ws in wb.worksheets:
        max_row, max_col = ws.max_row, ws.max_column
        if max_row < 1 or max_col < 1:
            continue
        for col in range(1, max_col + 1):
            letter = ws.cell(1, col).column_letter
            ws.column_dimensions[letter].width = WIDE_COLS.get(col, DEFAULT_WIDTH)
        for row in range(1, max_row + 1):
            ws.row_dimensions[row].height = ROW_HEIGHT
            is_header = row == 1
            for col in range(1, max_col + 1):
                cell = ws.cell(row, col)
                cell.alignment = Alignment(
                    wrap_text=True, vertical="center",
                    horizontal="center" if is_header else "left")
                cell.border = _BORDER
                if is_header:
                    cell.font = Font(bold=True, name=(cell.font.name or "Arial"),
                                     size=(cell.font.size or 10))
    out = out_path or path
    wb.save(out)
    return out


if __name__ == "__main__":
    import sys
    print("formatted:", format_workbook(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None))
