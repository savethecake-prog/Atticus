"""Source-driven field discovery.

The fields a product should carry come from inspecting the manufacturer's spec
sheet (and corroborating sources), not from whatever the template happens to
label. This reconciles a proposed field list against a template tab:

  - a field that matches an existing labelled spec column -> use that column
  - an unmatched field -> label a blank spare column inside the spec block
  - if no spare remains -> append a new column at the end and record it

reconcile() proposes; apply() writes the headers once the user confirms. Nothing
is added silently.
"""
from __future__ import annotations
import re
from openpyxl import load_workbook
from openpyxl.styles import Font
from common import BODY_FONT


def _toks(s):
    return set(re.findall(r"[a-z0-9]+", str(s or "").lower()))


_PLACEHOLDER = re.compile(r"^(spec|attribute|attr|field|feature|col|column)\s*\d+\s*:?\s*$", re.I)


def _is_spare(label):
    """A column is available to relabel if it is blank or a generic placeholder
    like 'Spec 1:' / 'Attribute 3' that carries no real field name."""
    if label in (None, ""):
        return True
    return bool(_PLACEHOLDER.match(str(label).strip()))


def _match_score(fkeys, hkeys):
    if not fkeys or not hkeys:
        return 0.0
    if fkeys == hkeys:
        return 1.0
    if (fkeys <= hkeys or hkeys <= fkeys) and min(len(fkeys), len(hkeys)) >= 2:
        return 0.9  # one is a genuine multi-word subset of the other
    return len(fkeys & hkeys) / len(fkeys | hkeys)  # Jaccard


def reconcile(tab_schema, discovered):
    """discovered: list of {"field": name, "keywords": [...]?}.
    Returns list of decisions: {field, action: existing|relabel|new, col|None}."""
    roles = tab_schema["roles"]
    labels = {int(c): v for c, v in tab_schema.get("labels", {}).items()}
    a, b = roles["spec_start"], roles["spec_end"]
    labelled = {c: labels.get(c) for c in range(a, b + 1)
                if labels.get(c) not in (None, "") and not _is_spare(labels.get(c))}
    spares = [c for c in range(a, b + 1) if _is_spare(labels.get(c))]
    decisions = []
    for d in discovered:
        field = d["field"] if isinstance(d, dict) else d
        keys = _toks(field)
        if isinstance(d, dict):
            for k in d.get("keywords", []):
                keys |= _toks(k)
        best_col, best = None, 0.0
        for c, lab in labelled.items():
            sc = _match_score(keys, _toks(lab))
            if sc > best:
                best_col, best = c, sc
        if best >= 0.6:
            header = labelled.get(best_col)
            exact = _toks(field) == _toks(header)
            decisions.append({"field": field, "action": "existing", "col": best_col,
                              "header": header, "exact": exact, "score": round(best, 2)})
        elif spares:
            decisions.append({"field": field, "action": "relabel", "col": spares.pop(0)})
        else:
            decisions.append({"field": field, "action": "new", "col": None})
    return decisions


def apply(template_path, schema, tab, decisions):
    """Write headers for relabel/new decisions; record new columns in the schema.
    Returns {field: column}."""
    wb = load_workbook(template_path)
    ws = wb[tab]
    t = schema["tabs"][tab]
    labels = t.setdefault("labels", {})
    extra = t.setdefault("extra_fields", {})
    notes = t.setdefault("field_notes", {})
    next_col = ws.max_column + 1
    field_col = {}
    for d in decisions:
        if d["action"] == "existing":
            field_col[d["field"]] = d["col"]
            if d.get("exact") is False:
                notes[str(d["col"])] = (f"column header '{d.get('header')}' filled from source field "
                                        f"'{d['field']}' - names not identical; confirm this is the right column")
        elif d["action"] == "relabel":
            c = d["col"]
            ws.cell(t["header_row"], c, value=d["field"]).font = Font(bold=True, **BODY_FONT)
            labels[str(c)] = d["field"]
            field_col[d["field"]] = c
        else:  # new
            c = next_col; next_col += 1
            ws.cell(t["header_row"], c, value=d["field"]).font = Font(bold=True, **BODY_FONT)
            labels[str(c)] = d["field"]
            extra[d["field"]] = c
            field_col[d["field"]] = c
    wb.save(template_path)
    return field_col


def from_spec_table(raw):
    """Turn a manufacturer spec table into a COMPLETE discovered-field list.

    The cause of thin listings is hand-picking a core subset of fields. This makes
    capturing the whole published table the easy path: paste the manufacturer's spec
    table and get one discovered field per row, in order, deduplicated. Feed the
    result straight to reconcile(). Accepts a dict {label: value}, a list of
    "Label: value" / "Label<tab>value" lines, or a multiline string.
    """
    if isinstance(raw, dict):
        rows = list(raw.keys())
    else:
        lines = raw.splitlines() if isinstance(raw, str) else list(raw)
        rows = []
        for ln in lines:
            s = str(ln).strip()
            if not s:
                continue
            if "\t" in s:
                rows.append(s.split("\t", 1)[0])
            elif ":" in s:
                rows.append(s.split(":", 1)[0])
            else:
                rows.append(s)  # a bare label with no value still counts as a field
    out, seen = [], set()
    for label in rows:
        lab = str(label).strip().rstrip(":").strip()
        key = re.sub(r"[^a-z0-9]+", " ", lab.lower()).strip()
        if not key or key in seen:
            continue
        seen.add(key)
        out.append({"field": lab})
    return out


if __name__ == "__main__":
    tab_schema = {"roles": {"spec_start": 3, "spec_end": 25, "header_row": 1},
                  "labels": {"3": "Electrical conductivity:", "4": "Supported application:",
                             "5": "Components:", "19": "Colour:", "20": "Operating temperature: "}}
    tab_schema["roles"]["header_row"] = 1
    discovered = [{"field": "Electrical conductivity"}, {"field": "Thermal conductivity (W/mK)"},
                  {"field": "Viscosity (Pa.s)"}, {"field": "Density (g/cm3)"}]
    dec = reconcile(tab_schema, discovered)
    by = {d["field"]: (d["action"], d["col"]) for d in dec}
    assert by["Electrical conductivity"][0] == "existing" and by["Electrical conductivity"][1] == 3, by
    assert by["Thermal conductivity (W/mK)"][0] == "relabel" and by["Thermal conductivity (W/mK)"][1] in (6, 7, 8), by
    assert by["Viscosity (Pa.s)"][0] == "relabel", by
    assert by["Density (g/cm3)"][0] == "relabel", by
    print("fields reconcile self-test passed:", by)
