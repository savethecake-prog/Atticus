"""Takealot (TAL) exporter for the salt-input-builder skill.

Implements the exporter contract (see root SKILL.md): consumes a verified job and produces
Takealot's official upload artifacts, preserving their format and enforcing their own rules.

Operates on the USER-PROVIDED live template at run time. Templates are never bundled
(they change); only a compact contract spec lives in assets/takealot-formats.json.

Status: routing + rule helpers are implemented and tested. The two artifact writers are
scaffolded with a fixed interface (next build increment).
"""
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

LOADSHEET_SHEET = "Loadsheet"
CATEGORY_SHEET = "Category Tree Lookup"
DATA_START_LOADSHEET = 7
DATA_START_EDIT_REQUEST = 5
# stable standard-column contract (1-indexed) on the loadsheet
STD_COLS = {1: "variant", 2: "sku", 3: "variant_code", 4: "main_category", 5: "lowest_category",
            6: "barcode", 7: "title", 8: "subtitle", 9: "description", 10: "whats_in_box",
            11: "brand", 12: "colour_main", 13: "colour_secondary", 14: "colour_name",
            15: "model_number", 16: "main_material"}


def route_product(tsin, category, trees):
    """Decide the artifact for one product.
    tsin: the product's TSIN (truthy if already listed).
    category: the product's category string/label from the job.
    trees: {'105': [..tree rows..], '120': [..]} from load_category_tree().
    Returns 'edit_request' | 'loadsheet_105' | 'loadsheet_120' | 'unrouted'.
    """
    if tsin not in (None, "", 0):
        return "edit_request"
    cat = _norm(category)
    best = None
    for key, rows in trees.items():
        for row in rows:
            hay = _norm(f"{row.get('main', '')} {row.get('lowest', '')}")
            if cat and (cat in hay or hay_overlap(cat, hay)):
                best = f"loadsheet_{key}"
                break
        if best:
            break
    return best or "unrouted"


def hay_overlap(cat, hay):
    ct = set(t for t in cat.split() if len(t) > 3)
    return bool(ct) and len(ct & set(hay.split())) >= max(1, len(ct) // 2)


def _norm(s):
    return " ".join(str(s or "").lower().replace("->", " ").replace("(", " ").replace(")", " ").split())


def load_category_tree(template_path):
    """Return [{'main','lowest'}] from a loadsheet's Category Tree Lookup tab."""
    wb = load_workbook(template_path, read_only=True, data_only=True)
    ws = wb[CATEGORY_SHEET]
    out = []
    for r in range(4, ws.max_row + 1):
        m = ws.cell(r, 1).value
        if m:
            out.append({"main": m, "lowest": ws.cell(r, 2).value})
    wb.close()
    return out


def load_rules(template_path, sheet=LOADSHEET_SHEET):
    """Read data validations into {col_index: {'type', 'max'|'list'}}.
    This is how Takealot's own rules are enforced - never hardcode a rulebook.
    """
    wb = load_workbook(template_path)
    ws = wb[sheet]
    rules = {}
    for dv in ws.data_validations.dataValidation:
        for rng in str(dv.sqref).split():
            letters = "".join(ch for ch in rng.split(":")[0] if ch.isalpha())
            try:
                ci = column_index_from_string(letters)
            except Exception:
                continue
            r = {"type": dv.type}
            if dv.type == "textLength":
                try:
                    r["max"] = int(dv.formula1)
                except Exception:
                    r["max"] = dv.formula1
            elif dv.type == "list":
                r["list"] = dv.formula1
            rules.setdefault(ci, r)
    wb.close()
    return rules


def enforce_text_length(value, maxlen):
    """Return (ok, value, flag). Never silently truncate: over-length is flagged, value kept."""
    if value is None or maxlen in (None, ""):
        return True, value, None
    n = len(str(value))
    if n > int(maxlen):
        return False, value, f"exceeds {maxlen} chars (is {n}); needs shortening before submission"
    return True, value, None


def check_dropdown(value, allowed):
    """Return (ok, flag). allowed: iterable of permitted values, or None to skip."""
    if value in (None, "") or not allowed:
        return True, None
    if str(value) in {str(a) for a in allowed}:
        return True, None
    return False, "value not in the column's dropdown list; will be rejected by Takealot"


# --- artifact writers: fixed interface, next build increment -------------------------------
def write_edit_request(job, template_path, out_path, salt_content=None):
    """Write one Edit Request row per TSIN product into a copy of the live template.
    Preserves header rows 1-4 and dropdown validations; writes data from row 5; only fills
    changed columns; attribute name/value pairs copied verbatim from the loadsheet labels.
    Returns a QA report (list of per-cell flags). NOT YET IMPLEMENTED."""
    raise NotImplementedError("write_edit_request: next build increment")


def write_loadsheet(job, template_path, cat_key, out_path, salt_content=None):
    """Write new-product rows into a copy of the live loadsheet (105 or 120).
    Opens with keep_vba=True, preserves rows 1-6 + all validations, writes from row 7,
    maps STD_COLS + sourced Attribute.* columns, enforces load_rules(), leaves blanks blank,
    never invents. Returns a QA report. NOT YET IMPLEMENTED."""
    raise NotImplementedError("write_loadsheet: next build increment")
